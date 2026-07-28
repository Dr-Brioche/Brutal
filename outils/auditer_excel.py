# -*- coding: utf-8 -*-
"""
AUDITER LE CLASSEUR D'ÉQUILIBRAGE  (Excel  <->  code du jeu)
===========================================================

    python3 outils/auditer_excel.py

Le classeur `docs/BRUTAL-items-et-cartes.xlsx` est la base d'équilibrage : c'est
lui qui décide des chiffres, et les `outils/importer_*.py` les recopient dans le
code. Mais rien ne garantissait jusqu'ici que les DEUX côtés parlent bien du
même jeu : une carte ajoutée dans le code et oubliée dans l'Excel, un objet
renommé d'un seul côté, un prix manquant… passaient inaperçus.

Ce script compare **tout le classeur** au **contenu réel du code** et liste :

  ✗ ERREURS   — incohérences franches (id inconnu, nom qui diverge, prix
                manquant, ligne/colonne fantôme…). À corriger.
  ⚠ ALERTES   — points à regarder mais qui peuvent être volontaires.

Comment il lit le code : il ne devine rien par expressions régulières, il
DEMANDE au jeu. Les fichiers `jeu/data/*.js` sont des modules ES sans DOM :
`node` peut donc les importer tels quels et recracher tout en JSON (cf.
`dumper_code()`). Ce qu'on compare est donc exactement ce que le jeu utilise.

Rien n'est modifié : c'est un contrôle, pas un importer. Code de sortie 1 s'il
reste au moins une erreur (pratique pour l'enchaîner à un autre outil).
"""

import json
import re
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant :  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).resolve().parent))

RACINE = Path(__file__).resolve().parent.parent
CLASSEUR = RACINE / "docs" / "BRUTAL-items-et-cartes.xlsx"

# ---------------------------------------------------------------------------
# 1) Le code du jeu, vu par le jeu lui-même
# ---------------------------------------------------------------------------

# Un module ES qui importe les données et les recrache en JSON. On garde les
# noms ANGLAIS (`_nomEn` n'existe qu'une fois la traduction installée ; ici,
# hors navigateur, `nom` EST déjà l'anglais puisque le code est écrit en anglais).
SCRIPT_NODE = r"""
const { ITEMS, SETS, RARETES } = await import("./jeu/data/items.js");
const { CARTES } = await import("./jeu/data/cartes.js");
const { ENNEMIS } = await import("./jeu/data/ennemis.js");
const { RECETTES } = await import("./jeu/data/recettes.js");
const { VALEUR_OBJET, VALEUR_RESSOURCE } = await import("./jeu/data/valeurs.js");
const { STATS_HEROS_BASE } = await import("./jeu/data/heros_base.js");
const { LOOTS_PROFONDEUR, CHANCES_RARETE_PROFONDEUR } = await import("./jeu/data/profondeur.js");

const sortie = {
  raretes: Object.fromEntries(Object.entries(RARETES).map(([k, v]) => [k, v.nom])),
  items: Object.fromEntries(Object.entries(ITEMS).map(([k, v]) => [k, {
    nom: v.nom, categorie: v.categorie, rarete: v.rarete,
    cartes: v.cartes ?? [], empilable: !!v.empilable,
    vitesseMinage: v.vitesseMinage ?? null, efficacite: v.efficacite ?? null,
  }])),
  cartes: Object.fromEntries(Object.entries(CARTES).map(([k, v]) => [k, {
    nom: v.nom, cout: v.cout, type: v.type, portee: v.portee ?? "melee",
    aoe: v.aoe === true, effets: (v.effets ?? []).map((e) => e.type),
  }])),
  sets: Object.fromEntries(Object.entries(SETS).map(([k, v]) => [k, {
    nom: v.nom, pieces: v.pieces,
  }])),
  ennemis: Object.fromEntries(ENNEMIS.map((e) => [e.id, {
    nom: e.nom, niveau: e.niveau, famille: e.famille, pv: e.pv,
    attaque: e.attaque, hits: e.attaqueHits ?? 1, xp: e.xp, vitesse: e.vitesse,
    grand: !!e.grand, actions: (e.actions ?? []).map((a) => a.type),
  }])),
  recettes: RECETTES.map((r) => r.resultat),
  valeurObjet: VALEUR_OBJET,
  valeurRessource: VALEUR_RESSOURCE,
  heros: STATS_HEROS_BASE,
  loots: Object.fromEntries(LOOTS_PROFONDEUR.map((l) => [l.id, {
    nom: l.nom, effet: l.effet, icone: l.icone, valeurs: l.valeurs,
  }])),
  chances: CHANCES_RARETE_PROFONDEUR,
};
process.stdout.write(JSON.stringify(sortie));
"""


def dumper_code():
    """Lance node sur les modules de données et renvoie tout le contenu du jeu."""
    r = subprocess.run(
        ["node", "--input-type=module", "-e", SCRIPT_NODE],
        cwd=RACINE, capture_output=True, text=True,
    )
    if r.returncode != 0:
        print("Impossible de lire les données du jeu avec node :\n" + r.stderr, file=sys.stderr)
        sys.exit(1)
    return json.loads(r.stdout)


# ---------------------------------------------------------------------------
# 2) Petits outils de lecture
# ---------------------------------------------------------------------------

def txt(cell):
    """Contenu d'une cellule, en texte propre ('' si vide)."""
    if cell is None:
        return ""
    return str(cell).strip()


def nombre(v):
    """Contenu d'une cellule en nombre, ou None si ce n'en est pas un."""
    s = txt(v).replace(",", ".").replace("%", "").strip()
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f == int(f) else f


def est_section(ligne):
    """Ligne de séparation (« ▸ ARMES », « ── DÉGÂTS ── »). Elle se reconnaît à
    sa puce, ou au fait que seule la colonne A est remplie."""
    a = txt(ligne[0])
    if not a:
        return False
    return a.startswith(("▸", "──", "—")) or not any(txt(c) for c in ligne[1:])


class Rapport:
    """Collecte les remarques, rangées par onglet, et les affiche à la fin."""

    def __init__(self):
        self.erreurs = []   # (onglet, ligne, message)
        self.alertes = []

    def erreur(self, onglet, ligne, message):
        self.erreurs.append((onglet, ligne, message))

    def alerte(self, onglet, ligne, message):
        self.alertes.append((onglet, ligne, message))

    def afficher(self):
        for titre, lot, marque in (("ERREURS", self.erreurs, "✗"),
                                   ("ALERTES (à regarder, pas forcément faux)", self.alertes, "⚠")):
            print(f"\n{'=' * 70}\n  {len(lot)} {titre}\n{'=' * 70}")
            onglet_courant = None
            for onglet, ligne, message in lot:
                if onglet != onglet_courant:
                    print(f"\n--- {onglet}")
                    onglet_courant = onglet
                ou = f"l.{ligne}" if ligne else "   "
                print(f"  {marque} {ou:>7}  {message}")
        if not self.erreurs:
            print("\n✔ Aucune erreur : le classeur et le code disent la même chose.")


# ---------------------------------------------------------------------------
# 3) Contrôles de forme (mise en page du classeur)
# ---------------------------------------------------------------------------

def auditer_forme(wb, rap):
    """Lignes et colonnes FANTÔMES : vides, mais comptées dans l'étendue de
    l'onglet. Elles font apparaître des cases inutiles, décalent les filtres et
    trompent tout script qui lit « jusqu'au bout ».

    Une colonne n'est PAS fantôme si une cellule FUSIONNÉE remplie l'atteint
    (un titre étalé sur toute la largeur, par exemple) — ce cas est normal.
    En revanche une fusion vide, elle, maintient des colonnes mortes : c'est
    exactement ce qui traînait dans l'onglet Monstres."""
    for ws in wb.worksheets:
        vides_bas = 0
        for r in range(ws.max_row, 0, -1):
            if any(txt(c.value) for c in ws[r]):
                break
            vides_bas += 1
        if vides_bas:
            rap.erreur(ws.title, None, f"{vides_bas} ligne(s) vide(s) en fin d'onglet")

        # Colonne la plus à droite réellement occupée : par une valeur, ou par
        # l'extension d'une fusion dont la cellule d'origine porte du texte.
        occupee = 0
        for ligne in ws.iter_rows():
            for cel in ligne:
                if txt(cel.value):
                    occupee = max(occupee, cel.column)
        for m in ws.merged_cells.ranges:
            if txt(ws.cell(row=m.min_row, column=m.min_col).value):
                occupee = max(occupee, m.max_col)
            else:
                rap.erreur(ws.title, m.min_row, f"fusion VIDE {m} — elle maintient des colonnes mortes")
        if ws.max_column > occupee:
            rap.erreur(ws.title, None, f"{ws.max_column - occupee} colonne(s) vide(s) à droite")


# ---------------------------------------------------------------------------
# 4) Contrôles par onglet
# ---------------------------------------------------------------------------

TYPES_CARTE = {"attaque": "Attaque", "defense": "Défense", "buff": "Buff"}
# « Portée » = attaque à distance (la Force ne s'applique pas), « Mêlée » = au corps à corps.
PORTEES = {"melee": "Mêlée", "range": "Portée"}
# Catégories d'items listées dans l'onglet « Items » (l'ÉQUIPEMENT). Les outils,
# ressources, trésors et parchemins de recette n'y ont rien à faire : ils ne
# donnent pas de cartes.
CATEGORIES_EQUIPEMENT = {"arme", "bouclier", "armure", "collier", "bague", "gant", "botte", "sac"}
# Ce qui se revend, donc ce qui a un prix dans l'onglet « Valeurs » : l'équipement,
# plus les trésors (objets faits pour être vendus) et l'outil (la pioche).
CATEGORIES_PRIX = CATEGORIES_EQUIPEMENT | {"tresor", "outil"}
# Effets qui touchent d'eux-mêmes TOUS les ennemis, sans passer par le drapeau
# `aoe: true` (lequel, lui, fait ré-appliquer toute la carte ennemi par ennemi).
EFFETS_TOUS_ENNEMIS = {"jugement-tous"}


def est_aoe(carte):
    """Une carte frappe-t-elle TOUS les ennemis d'un coup ? (≠ éclabousser ses
    voisins : un cleave n'est PAS de l'AOE.)"""
    return carte["aoe"] or any(e in EFFETS_TOUS_ENNEMIS for e in carte["effets"])


def effets_profondeur_reconnus():
    """La liste des effets de loot que le jeu sait appliquer. On la lit dans
    l'importer plutôt que de la recopier : une seule source, pas de dérive."""
    import importer_profondeur  # même dossier
    return importer_profondeur.EFFETS_OK


def calculer_donneurs(code):
    """Pour chaque carte, la liste [(id d'objet, combien de fois il la donne)],
    dans l'ordre des objets du catalogue. C'est l'inverse de `item.cartes`."""
    donneurs = {}
    for iid, it in code["items"].items():
        compte = {}
        for cid in it["cartes"]:
            compte[cid] = compte.get(cid, 0) + 1
        for cid, n in compte.items():
            donneurs.setdefault(cid, []).append((iid, n))
    return donneurs


def donnee_par(iid_liste, code):
    """Texte de la colonne « Donnée par » : les objets qui donnent la carte, avec
    « ×N » quand l'un d'eux la donne plusieurs fois."""
    morceaux = []
    for iid, n in iid_liste:
        nom = code["items"][iid]["nom"] if iid in code["items"] else iid
        morceaux.append(f"{nom} ×{n}" if n > 1 else nom)
    return ", ".join(morceaux)


def auditer_cartes(wb, code, rap):
    """Onglet « Cartes » : A Famille · B Nom · C ID · D Coût · E Type · F Rareté
    · G Portée · H AOE · I Donnée par · J Effet · K Visuel · L Cible (barème)."""
    ws = wb["Cartes"]
    O = "Cartes"
    donneurs = calculer_donneurs(code)
    vus = {}
    for i, ligne in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not any(txt(c) for c in ligne) or est_section(ligne):
            continue
        nom, cid, cout = txt(ligne[1]), txt(ligne[2]), nombre(ligne[3])
        typ, portee, aoe = txt(ligne[4]), txt(ligne[6]), txt(ligne[7]).lower()
        if not cid:
            rap.erreur(O, i, f"ligne sans ID (nom « {nom} »)")
            continue
        if cid in vus:
            rap.erreur(O, i, f"{cid} : ID en double (déjà ligne {vus[cid]})")
            continue
        vus[cid] = i
        c = code["cartes"].get(cid)
        if c is None:
            rap.erreur(O, i, f"{cid} : carte inconnue du code (jeu/data/cartes.js)")
            continue
        if nom != c["nom"]:
            rap.erreur(O, i, f"{cid} : nom « {nom} » ≠ code « {c['nom']} »")
        if cout is not None and cout != c["cout"]:
            rap.erreur(O, i, f"{cid} : coût {cout} ≠ code {c['cout']}")
        attendu = TYPES_CARTE.get(c["type"], c["type"])
        if typ and typ.lower() != attendu.lower():
            rap.erreur(O, i, f"{cid} : type « {typ} » ≠ code « {attendu }»")
        # Portée : « — » (tiret) pour les cartes qui ne visent personne.
        if portee and portee not in ("—", "-"):
            att = PORTEES.get(c["portee"], c["portee"])
            if portee.lower() != att.lower():
                rap.erreur(O, i, f"{cid} : portée « {portee} » ≠ code « {att} »")
        # AOE : la case dit « Oui » quand la carte frappe TOUS les ennemis d'un
        # coup. Une carte qui n'éclabousse que ses voisins directs (cleave,
        # contagion…) n'est PAS de l'AOE — sinon la colonne ne veut plus rien dire.
        if aoe and aoe != "oui":
            rap.erreur(O, i, f"{cid} : colonne AOE = « {txt(ligne[7])} » (écrire « Oui », ou laisser vide)")
        marque = aoe == "oui"
        if marque != est_aoe(c):
            quoi = "marqué AOE alors que le code ne touche que la cible / ses voisins" \
                if marque else "AOE dans le code, mais la case est vide"
            rap.erreur(O, i, f"{cid} : {quoi}")
        # « Donnée par » : la liste des objets qui donnent la carte. Regénérable —
        # elle doit coller exactement, sinon on ne sait plus qui débloque quoi.
        attendu = donnee_par(donneurs.get(cid, []), code)
        ecrit = txt(ligne[8])
        if not ecrit.startswith("—") and ecrit != attendu:
            rap.erreur(O, i, f"{cid} : « Donnée par » ≠ code — attendu « {attendu or '(personne)'} »")

    manquantes = [k for k in code["cartes"] if k not in vus]
    for k in manquantes:
        rap.erreur(O, None, f"carte ABSENTE du classeur : {k} ({code['cartes'][k]['nom']})")


def auditer_items(wb, code, rap):
    """Onglet « Items » : A Catégorie · B Nom · C ID · D Rareté · E→N Carte 1..10
    · O Effet hors-carte · P Total visé."""
    ws = wb["Items"]
    O = "Items"
    nom_vers_carte = {}
    for cid, c in code["cartes"].items():
        nom_vers_carte.setdefault(c["nom"], cid)
    vus = {}
    for i, ligne in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(txt(c) for c in ligne) or est_section(ligne):
            continue
        nom, iid, rarete = txt(ligne[1]), txt(ligne[2]), txt(ligne[3])
        if not iid:
            rap.erreur(O, i, f"ligne sans ID (nom « {nom} »)")
            continue
        if iid in vus:
            rap.erreur(O, i, f"{iid} : ID en double (déjà ligne {vus[iid]})")
            continue
        vus[iid] = i
        it = code["items"].get(iid)
        if it is None:
            rap.erreur(O, i, f"{iid} : objet inconnu du code (jeu/data/items.js)")
            continue
        if nom != it["nom"]:
            rap.erreur(O, i, f"{iid} : nom « {nom} » ≠ code « {it['nom']} »")
        att_rarete = code["raretes"].get(it["rarete"], it["rarete"])
        if rarete and rarete.lower() != att_rarete.lower():
            rap.erreur(O, i, f"{iid} : rareté « {rarete} » ≠ code « {att_rarete} »")

        # Une cellule par carte (colonnes E→N) : on compare les DEUX listes en
        # tenant compte des doublons (une arme peut donner 2× la même carte).
        cellules = [txt(c) for c in ligne[4:14] if txt(c)]
        trous = [j for j, c in enumerate(ligne[4:14]) if not txt(c)]
        if trous and any(j > min(trous) for j, c in enumerate(ligne[4:14]) if txt(c)):
            rap.erreur(O, i, f"{iid} : trou entre deux cartes (colonnes E→N doivent être remplies de gauche à droite)")
        attendues = sorted(code["cartes"][k]["nom"] if k in code["cartes"] else k for k in it["cartes"])
        if sorted(cellules) != attendues:
            manque = [n for n in attendues if attendues.count(n) > cellules.count(n)]
            trop = [n for n in cellules if cellules.count(n) > attendues.count(n)]
            detail = []
            if manque:
                detail.append("manque " + ", ".join(sorted(set(manque))))
            if trop:
                detail.append("en trop " + ", ".join(sorted(set(trop))))
            rap.erreur(O, i, f"{iid} : cartes ≠ code ({' · '.join(detail)})")
        for nc in cellules:
            if nc not in nom_vers_carte:
                rap.erreur(O, i, f"{iid} : carte « {nc} » introuvable dans le code")
        if len(it["cartes"]) > 10:
            rap.erreur(O, i, f"{iid} : {len(it['cartes'])} cartes — l'onglet n'a que 10 colonnes")

    for k, it in code["items"].items():
        if k not in vus and it["categorie"] in CATEGORIES_EQUIPEMENT:
            rap.erreur(O, None, f"objet ABSENT du classeur : {k} ({it['nom']})")


def auditer_sets(wb, code, rap):
    """Onglet « Sets » : A Set · B Pièces requises · C Déclencheur · D Bonus."""
    ws = wb["Sets"]
    O = "Sets"
    noms_items = {v["nom"]: k for k, v in code["items"].items()}
    vus = set()
    for i, ligne in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        nom = txt(ligne[0])
        if not nom:
            continue
        cle = next((k for k, v in code["sets"].items() if v["nom"] == nom), None)
        if cle is None:
            rap.erreur(O, i, f"set « {nom} » inconnu du code (jeu/data/items.js, SETS)")
            continue
        vus.add(cle)
        pieces = [p.strip() for p in txt(ligne[1]).split("+") if p.strip()]
        attendues = [code["items"][p]["nom"] for p in code["sets"][cle]["pieces"]]
        if sorted(pieces) != sorted(attendues):
            rap.erreur(O, i, f"{cle} : pièces « {' + '.join(pieces)} » ≠ code « {' + '.join(attendues)} »")
        for p in pieces:
            if p not in noms_items:
                rap.erreur(O, i, f"{cle} : pièce « {p} » introuvable dans le code")
        if not txt(ligne[2]) or not txt(ligne[3]):
            rap.erreur(O, i, f"{cle} : déclencheur ou bonus vide")
    for k, s in code["sets"].items():
        if k not in vus:
            rap.erreur(O, None, f"set ABSENT du classeur : {k} ({s['nom']})")


def auditer_effets(wb, code, rap):
    """Onglet « Effets » : A Identifiant · B Nom affiché · C Description · D Paramètres.
    Doit couvrir tous les types d'effet réellement employés par les cartes."""
    ws = wb["Effets"]
    O = "Effets"
    listes = set()
    for i, ligne in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        ident = txt(ligne[0])
        if not ident or ident.startswith("──"):
            continue
        listes.add(ident)
        if not txt(ligne[2]):
            rap.erreur(O, i, f"{ident} : description vide")
    employes = {t for c in code["cartes"].values() for t in c["effets"]}
    for t in sorted(employes - listes):
        rap.erreur(O, None, f"effet employé par une carte mais ABSENT du classeur : {t}")
    for t in sorted(listes - employes):
        rap.alerte(O, None, f"effet décrit dans le classeur mais employé par aucune carte : {t}")


def auditer_valeurs(wb, code, rap):
    """Onglet « Valeurs » : A id · B Nom · C Rareté · D Valeur. Un prix par objet
    non empilable (les ressources ont leur propre onglet)."""
    ws = wb["Valeurs"]
    O = "Valeurs"
    vus = {}
    for i, ligne in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        iid = txt(ligne[0])
        if not iid:
            continue
        if iid in vus:
            rap.erreur(O, i, f"{iid} : id en double (déjà ligne {vus[iid]})")
            continue
        vus[iid] = i
        it = code["items"].get(iid)
        if it is None:
            rap.erreur(O, i, f"{iid} : objet inconnu du code")
            continue
        if txt(ligne[1]) != it["nom"]:
            rap.erreur(O, i, f"{iid} : nom « {txt(ligne[1])} » ≠ code « {it['nom']} »")
        att = code["raretes"].get(it["rarete"], it["rarete"])
        if txt(ligne[2]).lower() != att.lower():
            rap.erreur(O, i, f"{iid} : rareté « {txt(ligne[2])} » ≠ code « {att} »")
        v = nombre(ligne[3])
        if v is None or v <= 0:
            rap.erreur(O, i, f"{iid} : valeur manquante ou nulle")
    for k, it in code["items"].items():
        if k in vus or it["categorie"] not in CATEGORIES_PRIX:
            continue
        rap.erreur(O, None, f"objet SANS PRIX : {k} ({it['nom']}, {it['categorie']})")

    # Rangement : l'onglet est trié par rareté croissante puis par nom. Un ajout
    # collé en fin de liste casse ce classement et rend la relecture pénible.
    rangs = {c: i for i, c in enumerate(code["raretes"])}
    ordre = [(rangs.get(code["items"][i]["rarete"], 99), code["items"][i]["nom"]) for i in vus]
    if ordre != sorted(ordre):
        rap.erreur(O, None, "l'onglet n'est plus trié par rareté puis par nom")


def auditer_ressources(wb, code, rap):
    """Onglet « Ressources » : A id · B Nom · C Rang · D Valeur."""
    ws = wb["Ressources"]
    O = "Ressources"
    vus = set()
    for i, ligne in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        rid = txt(ligne[0])
        if not rid:
            continue
        vus.add(rid)
        if rid not in code["items"]:
            rap.erreur(O, i, f"{rid} : ressource inconnue du code")
        v = nombre(ligne[3])
        if v is None or v <= 0:
            rap.erreur(O, i, f"{rid} : valeur manquante ou nulle")
    for k, v in code["valeurRessource"].items():
        if k not in vus:
            rap.erreur(O, None, f"ressource ABSENTE du classeur : {k}")


def auditer_recettes(wb, code, rap):
    """Onglet « Recettes » : un bloc par objet fabriqué. Ligne d'en-tête =
    A catégorie · B id · C nom · D rareté · E mot-clé ; puis la GRILLE (3 à 5
    lignes, colonnes B→F) avec les codes de ressources. Une ligne vide sépare
    deux blocs — c'est ce que lit outils/importer_recettes.py."""
    ws = wb["Recettes"]
    O = "Recettes"
    vus = {}
    for i, ligne in enumerate(ws.iter_rows(min_row=15, values_only=True), start=15):
        # En-tête de bloc : colonne A ET colonne B remplies (les lignes de grille
        # ont la colonne A vide).
        if not txt(ligne[0]) or not txt(ligne[1]):
            continue
        rid, nom, rarete = txt(ligne[1]), txt(ligne[2]), txt(ligne[3])
        if rid in vus:
            rap.erreur(O, i, f"{rid} : recette en double (déjà ligne {vus[rid]})")
            continue
        vus[rid] = i
        it = code["items"].get(rid)
        if it is None:
            rap.erreur(O, i, f"{rid} : objet fabriqué inconnu du code")
            continue
        if nom != it["nom"]:
            rap.erreur(O, i, f"{rid} : nom « {nom} » ≠ code « {it['nom']} »")
        if rarete and rarete != it["rarete"]:
            rap.erreur(O, i, f"{rid} : rareté « {rarete} » ≠ code « {it['rarete']} »")
        # La grille suit, jusqu'à la prochaine ligne vide. Elle ne doit RIEN
        # écrire en colonne A (l'importer y verrait un nouveau bloc) et elle doit
        # exister (un bloc sans grille = un objet qu'on ne peut pas fabriquer).
        cases = 0
        for j in range(i + 1, ws.max_row + 1):
            suivante = [c.value for c in ws[j]]
            if not any(txt(c) for c in suivante):
                break
            if txt(suivante[0]):
                break
            cases += sum(1 for c in suivante[1:6] if txt(c))
        if cases == 0:
            rap.erreur(O, i, f"{rid} : bloc sans grille d'ingrédients")
    for r in code["recettes"]:
        if r not in vus:
            rap.erreur(O, None, f"recette du code ABSENTE du classeur : {r}")


def auditer_profondeurs(wb, code, rap):
    """Onglets « Profondeurs » (A id · B Nom · C Effet · D Icône · E/F/G valeurs
    Normale/Rare/Épique) et « Profondeurs-chances » (rareté · %)."""
    ws = wb["Profondeurs"]
    O = "Profondeurs"
    # Les effets que le jeu sait appliquer. La liste vit dans l'importer : si on
    # écrit une phrase française dans la colonne Effet, l'import ne passe plus.
    effets_ok = effets_profondeur_reconnus()
    vus = set()
    for i, ligne in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        lid = txt(ligne[0])
        if not lid:
            continue
        vus.add(lid)
        l = code["loots"].get(lid)
        if l is None:
            rap.erreur(O, i, f"« {lid} » : loot inconnu du code — id mal orthographié, "
                             "ou import jamais relancé ?")
            continue
        if txt(ligne[1]) != l["nom"]:
            rap.erreur(O, i, f"{lid} : nom « {txt(ligne[1])} » ≠ code « {l['nom']} »")
        effet = txt(ligne[2])
        if effet not in effets_ok:
            rap.erreur(O, i, f"{lid} : effet « {effet or 'vide'} » non reconnu "
                             f"(attendus : {', '.join(sorted(effets_ok))})")
        elif effet != l["effet"]:
            rap.erreur(O, i, f"{lid} : effet « {effet} » ≠ code « {l['effet']} »")
        if not txt(ligne[3]).startswith("#"):
            rap.erreur(O, i, f"{lid} : icône « {txt(ligne[3]) or 'vide'} » — attendu une couleur hexa (#rrggbb)")
        for col, rarete in ((4, "normale"), (5, "rare"), (6, "epique")):
            brut = txt(ligne[col])
            attendu = l["valeurs"].get(rarete)
            valeur = None if brut in ("/", "—", "-", "") else nombre(brut)
            if brut and valeur is None and brut not in ("/", "—", "-"):
                rap.erreur(O, i, f"{lid} : valeur {rarete} « {brut} » illisible")
            elif valeur != attendu:
                rap.erreur(O, i, f"{lid} : {rarete} = {brut or 'vide'} ≠ code {attendu}")
    for k in code["loots"]:
        if k not in vus:
            rap.erreur(O, None, f"loot de profondeur ABSENT du classeur : {k}")

    ws = wb["Profondeurs-chances"]
    total = 0
    for i, ligne in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        r = txt(ligne[0])
        if not r:
            continue
        v = nombre(ligne[1])
        if v is None:
            rap.erreur("Profondeurs-chances", i, f"{r} : pourcentage illisible")
            continue
        total += v
        att = code["chances"].get(r)
        if att is not None and abs(att - v) > 0.001 and abs(att * 100 - v) > 0.001:
            rap.erreur("Profondeurs-chances", i, f"{r} : {v} % ≠ code {att}")
    if abs(total - 100) > 0.001:
        rap.erreur("Profondeurs-chances", None, f"la somme des chances fait {total} % (attendu 100 %)")


def auditer_monstres(wb, code, rap):
    """Onglet « Monstres » : A id · B nom · C niveau · D famille · E pv ·
    F attaque (« 4 » ou « 2x2 ») · G xp · H vitesse · I actions spéciales ·
    J grand · K zone de pop · L passif/note. Ligne 1 = en-têtes, ligne 2 = aide."""
    ws = wb["Monstres"]
    O = "Monstres"
    vus = {}
    for i, ligne in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        mid = txt(ligne[0])
        if not mid or est_section(ligne):
            continue
        # Sous le tableau vit le BARÈME DES RECETTES (des lignes de texte). Un id de
        # monstre est en minuscules-tirets : tout le reste n'est pas une ligne de
        # monstre, et n'a pas à être signalé comme « monstre inconnu ».
        if not re.fullmatch(r"[a-z0-9-]+", mid):
            continue
        if mid in vus:
            rap.erreur(O, i, f"{mid} : id en double (déjà ligne {vus[mid]})")
            continue
        vus[mid] = i
        m = code["ennemis"].get(mid)
        if m is None:
            rap.erreur(O, i, f"{mid} : monstre inconnu du code (jeu/data/ennemis.js)")
            continue
        if txt(ligne[1]) != m["nom"]:
            rap.erreur(O, i, f"{mid} : nom « {txt(ligne[1])} » ≠ code « {m['nom']} »")
        for col, cle, etiquette in ((2, "niveau", "niveau"), (4, "pv", "pv"),
                                    (6, "xp", "xp"), (7, "vitesse", "vitesse")):
            v = nombre(ligne[col])
            if v is None:
                rap.erreur(O, i, f"{mid} : {etiquette} illisible « {txt(ligne[col])} »")
            elif v != m[cle]:
                rap.erreur(O, i, f"{mid} : {etiquette} {v} ≠ code {m[cle]}")
        if txt(ligne[3]) != m["famille"]:
            rap.erreur(O, i, f"{mid} : famille « {txt(ligne[3])} » ≠ code « {m['famille']} »")
        # attaque : « 12 » (un coup) ou « 5x2 » (deux coups de 5)
        brut = txt(ligne[5]).lower().replace("×", "x")
        deg, hits = (brut.split("x") + ["1"])[:2] if "x" in brut else (brut, "1")
        if nombre(deg) != m["attaque"] or nombre(hits) != m["hits"]:
            rap.erreur(O, i, f"{mid} : attaque « {txt(ligne[5])} » ≠ code {m['attaque']}×{m['hits']}")
        grand = txt(ligne[9]) in ("1", "x", "oui")
        if grand != m["grand"]:
            rap.erreur(O, i, f"{mid} : colonne « grand » = {txt(ligne[9]) or 'vide'} ≠ code {m['grand']}")
        # Actions spéciales : UNE PAR LIGNE dans la cellule, avec son « (n%) ».
        # L'attaque de base compte pour une ligne (elle occupe le % restant). On
        # ne compare pas le texte (Brioche l'écrit librement) mais le NOMBRE :
        # une ligne oubliée par l'importer se voit tout de suite ici.
        lignes_act = [l for l in txt(ligne[8]).splitlines() if l.strip()]
        if not m["actions"] and lignes_act:
            rap.alerte(O, i, f"{mid} : {len(lignes_act)} action(s) écrite(s), aucune dans le code "
                             "(rencontre spéciale traitée à part ?)")
        elif len(lignes_act) != len(m["actions"]):
            rap.erreur(O, i, f"{mid} : {len(lignes_act)} action(s) écrite(s) ≠ {len(m['actions'])} dans le code")
        if not txt(ligne[10]):
            rap.alerte(O, i, f"{mid} : zone de pop non renseignée")
    for k, m in code["ennemis"].items():
        if k not in vus:
            rap.erreur(O, None, f"monstre ABSENT du classeur : {k} ({m['nom']})")


def auditer_heros(wb, code, rap):
    """Onglet « Héros » : A clé · B valeur · C description. Les clés doivent
    exister dans STATS_HEROS_BASE (l'importer accepte quelques synonymes)."""
    ws = wb["Héros"]
    O = "Héros"
    synonymes = {"agilite": "vitesseCombat", "agilité": "vitesseCombat"}
    vus = set()
    for i, ligne in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cle = txt(ligne[0])
        if not cle:
            continue
        reelle = synonymes.get(cle.lower(), cle)
        vus.add(reelle)
        if reelle not in code["heros"]:
            rap.erreur(O, i, f"{cle} : clé inconnue de STATS_HEROS_BASE")
            continue
        v = nombre(ligne[1])
        if v is None:
            rap.erreur(O, i, f"{cle} : valeur illisible")
        elif v != code["heros"][reelle]:
            rap.erreur(O, i, f"{cle} : {v} ≠ code {code['heros'][reelle]}")
        if not txt(ligne[2]):
            rap.alerte(O, i, f"{cle} : pas d'explication")
    for k in code["heros"]:
        if k not in vus:
            rap.erreur(O, None, f"stat de héros ABSENTE du classeur : {k}")


def auditer_pioches(wb, code, rap):
    """Onglet « Pioches » : A id · B Nom · C Rareté · D Vitesse minage · E Efficacité.
    Les trois chiffres doivent coller au code (bloc PIOCHES-AUTO d'items.js), sinon
    c'est que l'import n'a pas été relancé après une modification du classeur."""
    ws = wb["Pioches"]
    O = "Pioches"
    vus = set()
    for i, ligne in enumerate(ws.iter_rows(values_only=True), start=1):
        pid = txt(ligne[0] if ligne else None)
        if not pid or pid == "id" or pid not in code["items"]:
            continue
        vus.add(pid)
        it = code["items"][pid]
        if it["categorie"] != "outil":
            rap.erreur(O, i, f"{pid} : ce n'est pas un outil dans le code (catégorie « {it['categorie']} »)")
        if txt(ligne[1]) != it["nom"]:
            rap.erreur(O, i, f"{pid} : nom « {txt(ligne[1])} » ≠ code « {it['nom']} »")
        if txt(ligne[2]) != it["rarete"]:
            rap.erreur(O, i, f"{pid} : rareté « {txt(ligne[2])} » ≠ code « {it['rarete']} »")
        for col, cle, nom in ((3, "vitesseMinage", "vitesse de minage"), (4, "efficacite", "efficacité")):
            v = nombre(ligne[col])
            if v is None:
                rap.erreur(O, i, f"{pid} : {nom} illisible")
            elif v != it[cle]:
                rap.erreur(O, i, f"{pid} : {nom} {v} ≠ code {it[cle]} (import à relancer ?)")
    for k, it in code["items"].items():
        if it["categorie"] == "outil" and k not in vus:
            rap.erreur(O, None, f"pioche ABSENTE du classeur : {k} ({it['nom']})")


def auditer_general(wb, rap):
    """Onglet « Général » : A Domaine · B Réglage · C Valeur · D Explication ·
    E Où (fichier). Le fichier cité doit exister — sinon Brioche cherche dans
    le vide, et c'est le seul onglet sans importer (report à la main)."""
    ws = wb["Général"]
    O = "Général"
    for i, ligne in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        if not any(txt(c) for c in ligne) or est_section(ligne):
            continue
        reglage = txt(ligne[1])
        if not reglage:
            continue
        if not txt(ligne[2]):
            rap.erreur(O, i, f"« {reglage} » : valeur vide")
        if not txt(ligne[3]):
            rap.erreur(O, i, f"« {reglage} » : explication vide")
        fichier = txt(ligne[4])
        if not fichier:
            rap.erreur(O, i, f"« {reglage} » : colonne « Où (fichier) » vide")
            continue
        for f in [x.strip() for x in fichier.replace("+", ",").split(",") if x.strip()]:
            if not (RACINE / "jeu" / f).exists() and not (RACINE / f).exists():
                rap.erreur(O, i, f"« {reglage} » : fichier introuvable — {f}")


# ---------------------------------------------------------------------------

def main():
    if not CLASSEUR.exists():
        print(f"Classeur introuvable : {CLASSEUR}", file=sys.stderr)
        sys.exit(1)
    code = dumper_code()
    wb = load_workbook(CLASSEUR)
    rap = Rapport()

    attendus = ["Lisez-moi", "Cartes", "Items", "Pioches", "Sets", "Effets", "Valeurs", "Ressources",
                "Recettes", "Profondeurs", "Profondeurs-chances", "Monstres", "Héros", "Général"]
    for t in attendus:
        if t not in wb.sheetnames:
            rap.erreur("(classeur)", None, f"onglet manquant : {t}")
    for t in wb.sheetnames:
        if t not in attendus:
            rap.alerte("(classeur)", None, f"onglet inattendu : {t}")

    auditer_forme(wb, rap)
    auditer_cartes(wb, code, rap)
    auditer_items(wb, code, rap)
    auditer_sets(wb, code, rap)
    auditer_effets(wb, code, rap)
    auditer_valeurs(wb, code, rap)
    auditer_ressources(wb, code, rap)
    auditer_recettes(wb, code, rap)
    auditer_profondeurs(wb, code, rap)
    auditer_monstres(wb, code, rap)
    auditer_heros(wb, code, rap)
    auditer_pioches(wb, code, rap)
    auditer_general(wb, rap)

    print(f"Audit de {CLASSEUR.relative_to(RACINE)}")
    print(f"  {len(code['items'])} objets · {len(code['cartes'])} cartes · {len(code['sets'])} sets"
          f" · {len(code['ennemis'])} monstres · {len(code['recettes'])} recettes")
    rap.afficher()
    sys.exit(1 if rap.erreurs else 0)


if __name__ == "__main__":
    main()
