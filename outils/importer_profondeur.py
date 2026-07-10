# -*- coding: utf-8 -*-
"""
IMPORTER LES LOOTS DE PROFONDEUR  (Excel  ->  jeu/data/profondeur.js)
====================================================================

Les buffs de « run » qu'on ramasse en descendant dans les profondeurs sont
définis dans le classeur — la SOURCE :

    docs/BRUTAL-items-et-cartes.xlsx
      ->  onglet « Profondeurs »          : la table des loots (valeurs par rareté)
      ->  onglet « Profondeurs-chances »  : la chance de tirer chaque rareté

Puis on régénère la table lue par le jeu :

    python3 outils/importer_profondeur.py

Le script réécrit UNIQUEMENT le bloc balisé de jeu/data/profondeur.js :

    // <<PROFONDEUR-AUTO>>   ...généré...   // <<FIN-PROFONDEUR-AUTO>>

Onglet « Profondeurs » (en-tête ligne 6, données ensuite) :
    A = id · B = Nom · C = Effet · D = Icône (hex) · E = Normale · F = Rare · G = Épique
Effets reconnus (ils doivent matcher les champs du run, cf. systems/profondeur.js) :
    force · gold · celerite · armure · porte (garantit une porte de descente)
Onglet « Profondeurs-chances » :
    A = rarete (normale|rare|epique) · B = chance en %  (doit sommer à 100)
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant :  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RACINE = Path(__file__).resolve().parent.parent
XLSX = RACINE / "docs" / "BRUTAL-items-et-cartes.xlsx"
PROF_JS = RACINE / "jeu" / "data" / "profondeur.js"

DEBUT = "// <<PROFONDEUR-AUTO>>"
FIN = "// <<FIN-PROFONDEUR-AUTO>>"

EFFETS_OK = {"force", "gold", "celerite", "armure", "porte"}
RARETES = ("normale", "rare", "epique")


def lire_loots(wb):
    if "Profondeurs" not in wb.sheetnames:
        raise SystemExit("L'onglet « Profondeurs » est introuvable dans le classeur.")
    ws = wb["Profondeurs"]
    loots, vus = [], set()
    for i, lg in enumerate(ws.iter_rows(values_only=True), start=1):
        if not lg:
            continue
        iid = str(lg[0]).strip() if lg[0] else ""
        if iid == "id" or not re.fullmatch(r"[a-z0-9-]+", iid):
            continue  # titre / en-tête / ligne vide
        if iid in vus:
            raise SystemExit(f"« Profondeurs » : id en double « {iid} ».")
        vus.add(iid)
        nom = str(lg[1]).strip() if len(lg) > 1 and lg[1] else iid
        effet = str(lg[2]).strip() if len(lg) > 2 and lg[2] else ""
        if effet not in EFFETS_OK:
            raise SystemExit(f"« Profondeurs » ligne {i} : effet inconnu « {effet} » pour « {iid} » "
                             f"(attendus : {', '.join(sorted(EFFETS_OK))}).")
        icone = str(lg[3]).strip() if len(lg) > 3 and lg[3] else "#888888"
        try:
            vals = [int(round(float(lg[4 + k]))) for k in range(3)]  # E, F, G
        except (TypeError, ValueError, IndexError):
            raise SystemExit(f"« Profondeurs » ligne {i} : valeurs Normale/Rare/Épique illisibles pour « {iid} ».")
        loots.append((iid, nom, effet, icone, vals))
    if not loots:
        raise SystemExit("Aucun loot trouvé dans l'onglet « Profondeurs ».")
    return loots


def lire_chances(wb):
    if "Profondeurs-chances" not in wb.sheetnames:
        raise SystemExit("L'onglet « Profondeurs-chances » est introuvable dans le classeur.")
    ws = wb["Profondeurs-chances"]
    ch = {}
    for lg in ws.iter_rows(values_only=True):
        if not lg:
            continue
        r = str(lg[0]).strip().lower() if lg[0] else ""
        if r not in RARETES:
            continue
        try:
            ch[r] = float(lg[1])
        except (TypeError, ValueError, IndexError):
            raise SystemExit(f"« Profondeurs-chances » : chance illisible pour « {r} ».")
    manquants = [r for r in RARETES if r not in ch]
    if manquants:
        raise SystemExit(f"« Profondeurs-chances » : rareté(s) manquante(s) : {', '.join(manquants)}.")
    total = sum(ch.values())
    if abs(total - 100) > 0.01:
        raise SystemExit(f"« Profondeurs-chances » : les chances font {total} %, pas 100 %.")
    return ch


def bloc_js(loots, chances):
    lignes = ["export const LOOTS_PROFONDEUR = ["]
    for iid, nom, effet, icone, vals in loots:
        lignes.append(
            f'  {{ id: "{iid}", nom: "{nom}", effet: "{effet}", icone: "{icone}", '
            f'valeurs: {{ normale: {vals[0]}, rare: {vals[1]}, epique: {vals[2]} }} }},'
        )
    lignes.append("];")
    lignes.append("")
    lignes.append("// Chance de tirer chaque rareté à CHAQUE choix (somme = 1).")
    lignes.append("export const CHANCES_RARETE_PROFONDEUR = {")
    for r in RARETES:
        lignes.append(f"  {r}: {round(chances[r] / 100, 4)},")
    lignes.append("};")
    return "\n".join(lignes) + "\n"


def main():
    wb = load_workbook(XLSX, data_only=True)
    loots = lire_loots(wb)
    chances = lire_chances(wb)
    corps = bloc_js(loots, chances)

    contenu = PROF_JS.read_text(encoding="utf-8") if PROF_JS.exists() else ""
    if DEBUT not in contenu or FIN not in contenu:
        contenu = (
            "// Loots de PROFONDEUR (buffs de « run ») régénérés par\n"
            "// outils/importer_profondeur.py depuis le classeur Excel\n"
            "// (onglets « Profondeurs » = table, « Profondeurs-chances » = probabilités).\n"
            "// NE PAS éditer le bloc auto à la main — passer par l'Excel.\n\n"
            f"{DEBUT}\n{FIN}\n"
        )
    genere = f"{DEBUT}\n" + corps + FIN
    avant = contenu.split(DEBUT)[0]
    apres = contenu.split(FIN)[1]
    PROF_JS.write_text(avant + genere + apres, encoding="utf-8")
    print(f"OK — {len(loots)} loots + chances {chances} écrits dans {PROF_JS.relative_to(RACINE)}.")


if __name__ == "__main__":
    main()
