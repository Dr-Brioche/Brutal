# -*- coding: utf-8 -*-
"""
IMPORTER LA VALEUR DES OBJETS  (Excel  ->  jeu/data/valeurs.js)
==============================================================

Brioche fixe la VALEUR de chaque objet (sa valeur de référence, ≈ prix d'achat)
dans le classeur — la SOURCE :

    docs/BRUTAL-items-et-cartes.xlsx  ->  onglet « Valeurs »

Puis on lance ce script pour régénérer la table lue par le jeu :

    python3 outils/importer_valeurs.py

Le script réécrit UNIQUEMENT le bloc balisé de jeu/data/valeurs.js :

    // <<VALEURS-AUTO>>   ...généré...   // <<FIN-VALEURS-AUTO>>

RÈGLE (cf. docs/concept.md « Valeur des objets ») : la valeur est LIBRE (Brioche
l'ajuste à la main), mais elle a été semée autour d'une cible par rareté
(Common 400 · Uncommon 2 000 · Rare 10 000 · Epic 80 000 · Legendary 250 000)
avec ± 20 % de variation pour que deux objets ne vaillent pas pile pareil. Le
PRIX DE VENTE (au marchand / à l'HV) en découle automatiquement dans le jeu
(−20 à −30 %) ; ici on ne touche QUE la valeur de référence.

Format de l'onglet « Valeurs » : une ligne d'en-tête, puis une ligne par objet :
    colonne A = id · B = nom (info) · C = rareté (info) · D = VALEUR (éditable)
Seules les colonnes A (id) et D (valeur) sont lues ; B et C sont décoratives.
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant :  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RACINE = Path(__file__).resolve().parent.parent
XLSX = RACINE / "docs" / "BRUTAL-items-et-cartes.xlsx"
VALEURS_JS = RACINE / "jeu" / "data" / "valeurs.js"
ITEMS_JS = RACINE / "jeu" / "data" / "items.js"

DEBUT = "// <<VALEURS-AUTO>>"
FIN = "// <<FIN-VALEURS-AUTO>>"


def ids_items_connus():
    txt = ITEMS_JS.read_text(encoding="utf-8")
    return set(re.findall(r'"([a-z0-9-]+)":\s*\{\s*id:', txt))


def lire_onglet(wb, nom, obligatoire=True):
    """Lit un onglet id|…|…|valeur → liste [(id, valeur)]. Colonne A = id
    (minuscules/chiffres/tirets), colonne D = valeur. Titres/en-têtes ignorés."""
    if nom not in wb.sheetnames:
        if obligatoire:
            raise SystemExit(f"L'onglet « {nom} » est introuvable dans le classeur.")
        return []
    ws = wb[nom]
    valeurs = []
    for i, ligne in enumerate(ws.iter_rows(values_only=True)):
        if not ligne:
            continue
        iid = str(ligne[0]).strip() if ligne[0] else ""
        if iid == "id" or not re.fullmatch(r"[a-z0-9-]+", iid):
            continue  # titre / en-tête / ligne vide
        brut = ligne[3] if len(ligne) > 3 else None  # colonne D
        if brut is None or str(brut).strip() == "":
            raise SystemExit(f"« {nom} » ligne {i + 1} : « {iid} » n'a pas de valeur (colonne D).")
        try:
            val = int(round(float(brut)))
        except (TypeError, ValueError):
            raise SystemExit(f"« {nom} » ligne {i + 1} : valeur illisible pour « {iid} » : {brut!r}.")
        if val < 1:
            raise SystemExit(f"« {nom} » ligne {i + 1} : valeur < 1 pour « {iid} ».")
        valeurs.append((iid, val))
    return valeurs


def bloc_js(nom_const, valeurs, connus, nom_onglet):
    """Construit `export const <nom_const> = { … };` en validant ids + doublons."""
    vus = set()
    lignes = []
    for iid, val in valeurs:
        if iid in vus:
            raise SystemExit(f"Doublon d'id dans l'onglet « {nom_onglet} » : « {iid} ».")
        vus.add(iid)
        if iid not in connus:
            print(f"  ! Attention : « {iid} » n'existe pas dans items.js (faute de frappe ?)",
                  file=sys.stderr)
        lignes.append(f'  "{iid}": {val},')
    return f"export const {nom_const} = {{\n" + "\n".join(lignes) + "\n};\n"


def main():
    connus = ids_items_connus()
    wb = load_workbook(XLSX, data_only=True)
    objets = lire_onglet(wb, "Valeurs", obligatoire=True)
    ressources = lire_onglet(wb, "Ressources", obligatoire=False)
    if not objets:
        raise SystemExit("Aucune valeur trouvée dans l'onglet « Valeurs ».")

    corps = bloc_js("VALEUR_OBJET", objets, connus, "Valeurs")
    if ressources:
        corps += "\n// Valeur de BASE (prix marché) des RESSOURCES — onglet « Ressources ».\n"
        corps += bloc_js("VALEUR_RESSOURCE", ressources, connus, "Ressources")
    else:
        corps += "\nexport const VALEUR_RESSOURCE = {};\n"

    contenu = VALEURS_JS.read_text(encoding="utf-8") if VALEURS_JS.exists() else ""
    if DEBUT not in contenu or FIN not in contenu:
        contenu = (
            "// Valeurs ÉDITABLES régénérées par outils/importer_valeurs.py depuis le\n"
            "// classeur Excel (onglets « Valeurs » = objets, « Ressources » = minerais/\n"
            "// bois/cuir). NE PAS éditer le bloc auto à la main.\n"
            "//  • VALEUR_OBJET    : valeur de référence d'un objet (cf. data/items.js) ;\n"
            "//    le prix de vente en découle (−20 à −30 %).\n"
            "//  • VALEUR_RESSOURCE : prix de base marché d'une ressource (cf. marche.js).\n\n"
            f"{DEBUT}\n{FIN}\n"
        )

    genere = f"{DEBUT}\n" + corps + FIN
    avant = contenu.split(DEBUT)[0]
    apres = contenu.split(FIN)[1]
    VALEURS_JS.write_text(avant + genere + apres, encoding="utf-8")
    print(f"OK — {len(objets)} objets" + (f" + {len(ressources)} ressources" if ressources else "")
          + f" écrits dans {VALEURS_JS.relative_to(RACINE)}.")


if __name__ == "__main__":
    main()
