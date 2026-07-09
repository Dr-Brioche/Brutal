# -*- coding: utf-8 -*-
"""
IMPORTER LES RECETTES DE CRAFT  (Excel  ->  jeu/data/recettes.js)
=================================================================

Brioche édite les patterns de craft dans le classeur Excel (la SOURCE) :

    docs/BRUTAL-items-et-cartes.xlsx  ->  onglet « Recettes »

Puis on lance ce script UNE FOIS pour régénérer le tableau RECETTES du jeu :

    python3 outils/importer_recettes.py

Le script réécrit UNIQUEMENT le bloc balisé dans jeu/data/recettes.js :

    // <<RECETTES-AUTO>>   ...généré...   // <<FIN-RECETTES-AUTO>>

Le reste du fichier (les QUALITES de forge, les commentaires) n'est jamais touché.

----------------------------------------------------------------------------
FORMAT DE L'ONGLET « Recettes »
----------------------------------------------------------------------------
Le classeur se lit par BLOCS. Un bloc = une arme :

    colonne A = "ARME"  ->  B = id de l'arme, C = nom, D = rareté, E = thème
    puis les lignes SUIVANTES (colonne A vide) donnent la GRILLE du motif,
    dans les colonnes B..F (5 cases de large max, 5 lignes de haut max).

Chaque case de la grille contient un CODE de 2 lettres (voir CODE_MINERAI)
ou reste vide. C'est la FORME qui compte (façon Minecraft) : la position
absolue sur la table n'a aucune importance, seule la forme relative compte.

Le bloc se termine à la ligne vide suivante (ou au prochain "ARME").
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant :  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RACINE = Path(__file__).resolve().parent.parent
XLSX = RACINE / "docs" / "BRUTAL-items-et-cartes.xlsx"
RECETTES_JS = RACINE / "jeu" / "data" / "recettes.js"
ITEMS_JS = RACINE / "jeu" / "data" / "items.js"

# Code Excel (2 lettres) -> id de ressource (cf. jeu/data/items.js)
CODE_MINERAI = {
    "Fe": "fer",       "Cu": "cuivre",   "Co": "charbon",  "St": "pierre-taillee",
    "Wd": "bois",      "Ag": "argent",   "Au": "or",       "Ma": "malachite",
    "La": "lapis",     "Am": "amethyste", "Ti": "titane",  "Em": "emeraude",
    "Ru": "rubis",     "Sa": "saphir",   "Di": "diamant",  "Mi": "mithril",
    "On": "onyx",      "Su": "pierre-solaire",
    # Paliers de bois et de cuir (uncommon / rare) :
    "Ws": "bois-sombre",   "We": "bois-enchante",
    "Cr": "cuir",          "Ce": "cuir-epais",   "Cx": "cuir-etrange",
}

# id de ressource -> lettre unique utilisée comme clé de légende dans recettes.js
ID_CARACTERE = {
    "fer": "F", "cuivre": "C", "charbon": "K", "pierre-taillee": "S", "bois": "W",
    "argent": "A", "or": "O", "malachite": "M", "lapis": "L", "amethyste": "Y",
    "titane": "T", "emeraude": "E", "rubis": "R", "saphir": "P", "diamant": "D",
    "mithril": "I", "onyx": "X", "pierre-solaire": "U",
    "bois-sombre": "B", "bois-enchante": "N",
    "cuir": "H", "cuir-epais": "J", "cuir-etrange": "Q",
}

DEBUT = "// <<RECETTES-AUTO>>"
FIN = "// <<FIN-RECETTES-AUTO>>"


def ids_items_connus():
    """Tous les ids déclarés dans items.js (pour repérer les fautes de frappe)."""
    txt = ITEMS_JS.read_text(encoding="utf-8")
    return set(re.findall(r'"([a-z0-9-]+)":\s*\{\s*id:', txt))


def lire_recettes(chemin):
    wb = load_workbook(chemin, data_only=True)
    if "Recettes" not in wb.sheetnames:
        raise SystemExit("L'onglet « Recettes » est introuvable dans le classeur.")
    ws = wb["Recettes"]
    lignes = list(ws.iter_rows(values_only=True))

    # Une ligne est un EN-TÊTE de recette si : colonne A = catégorie (non vide,
    # ex. « Arme », « Armure »…) ET colonne B = un id valide (minuscules/tirets).
    # → la colonne A range chaque objet dans sa catégorie sans casser la lecture.
    def est_entete(lg):
        colA = str(lg[0]).strip() if lg and lg[0] else ""
        colB = str(lg[1]).strip() if lg and len(lg) > 1 and lg[1] else ""
        return bool(colA) and bool(re.fullmatch(r"[a-z0-9-]+", colB))

    recettes = []
    i = 0
    while i < len(lignes):
        ligne = lignes[i]
        if not est_entete(ligne):
            i += 1
            continue
        resultat = str(ligne[1]).strip()

        # Lire la grille : lignes suivantes (cols B..F), jusqu'à la prochaine
        # en-tête de recette ou une ligne vide.
        grille = []
        j = i + 1
        while j < len(lignes):
            lg = lignes[j]
            if est_entete(lg):
                break
            cases = [lg[c] if lg and len(lg) > c else None for c in range(1, 6)]  # B..F
            if all(v is None or str(v).strip() == "" for v in cases):
                break  # ligne vide -> fin du bloc
            grille.append([str(v).strip() if v is not None else "" for v in cases])
            j += 1
            if len(grille) >= 5:
                break

        recettes.append({"resultat": resultat, "grille": grille})
        i = j
    return recettes


def grille_vers_ids(recette):
    """Convertit la grille de codes en grille d'ids (ou None) + relève les ingrédients."""
    out = []
    for rang in recette["grille"]:
        ligne_ids = []
        for code in rang:
            if not code:
                ligne_ids.append(None)
                continue
            if code not in CODE_MINERAI:
                raise SystemExit(
                    f"Recette « {recette['resultat']} » : code inconnu « {code} » "
                    f"(codes valides : {', '.join(sorted(CODE_MINERAI))})."
                )
            ligne_ids.append(CODE_MINERAI[code])
        out.append(ligne_ids)
    return out


def recadrer(grille_ids):
    """Réduit la grille à sa boîte englobante (comme systems/craft.js)."""
    lignes_pleines = [r for r in range(len(grille_ids)) if any(grille_ids[r])]
    if not lignes_pleines:
        return []
    largeur = max((len(r) for r in grille_ids), default=0)
    cols_pleines = [c for c in range(largeur)
                    if any(c < len(grille_ids[r]) and grille_ids[r][c] for r in range(len(grille_ids)))]
    r0, r1 = min(lignes_pleines), max(lignes_pleines)
    c0, c1 = min(cols_pleines), max(cols_pleines)
    return [[(grille_ids[r][c] if c < len(grille_ids[r]) else None)
             for c in range(c0, c1 + 1)] for r in range(r0, r1 + 1)]


def forme_et_legende(grille_ids):
    """Grille d'ids -> (forme=lignes de lettres, legende=lettre->id)."""
    recadree = recadrer(grille_ids)
    legende = {}
    forme = []
    for rang in recadree:
        ligne = ""
        for cellule in rang:
            if cellule is None:
                ligne += "."
            else:
                car = ID_CARACTERE.get(cellule)
                if car is None:
                    raise SystemExit(f"Aucune lettre de légende prévue pour « {cellule} ».")
                legende[car] = cellule
                ligne += car
        forme.append(ligne)
    return forme, legende, recadree


def js_recette(resultat, forme, legende):
    lignes_forme = ",\n".join(f'      "{l}"' for l in forme)
    entrees_leg = ", ".join(f'{k}: "{v}"' for k, v in legende.items())
    return (
        "  {\n"
        f'    resultat: "{resultat}",\n'
        "    forme: [\n"
        f"{lignes_forme},\n"
        "    ],\n"
        f"    legende: {{ {entrees_leg} }},\n"
        "  },"
    )


def main():
    connus = ids_items_connus()
    recettes = lire_recettes(XLSX)
    if not recettes:
        raise SystemExit("Aucune recette trouvée dans l'onglet « Recettes ».")

    blocs = []
    motifs_vus = []  # (resultat, grille recadrée) pour détecter les collisions
    for rec in recettes:
        resultat = rec["resultat"]
        if resultat not in connus:
            print(f"  ! Attention : « {resultat} » n'existe pas dans items.js "
                  f"(faute de frappe ?)", file=sys.stderr)
        grille_ids = grille_vers_ids(rec)
        if not recadrer(grille_ids):
            raise SystemExit(f"Recette « {resultat} » : grille vide.")
        forme, legende, recadree = forme_et_legende(grille_ids)

        # Unicité : deux recettes ne doivent JAMAIS avoir le même motif recadré
        # (sinon la forge ne saurait pas laquelle fabriquer — cf. systems/craft.js).
        for autre_res, autre_motif in motifs_vus:
            if autre_motif == recadree:
                raise SystemExit(
                    f"COLLISION de motif : « {resultat} » et « {autre_res} » "
                    f"ont exactement la même forme + les mêmes ingrédients."
                )
        motifs_vus.append((resultat, recadree))
        blocs.append(js_recette(resultat, forme, legende))

    contenu = RECETTES_JS.read_text(encoding="utf-8")
    if DEBUT not in contenu or FIN not in contenu:
        raise SystemExit(
            f"Les balises {DEBUT} / {FIN} sont absentes de recettes.js. "
            "Ajoute-les autour du tableau RECETTES avant de relancer."
        )

    genere = (
        f"{DEBUT}\n"
        "// Bloc GÉNÉRÉ automatiquement par outils/importer_recettes.py — NE PAS\n"
        "// éditer à la main : modifier l'onglet « Recettes » du classeur Excel puis\n"
        "// relancer le script. (La SOURCE, c'est le classeur.)\n"
        "export const RECETTES = [\n"
        + "\n".join(blocs)
        + "\n];\n"
        f"{FIN}"
    )

    avant = contenu.split(DEBUT)[0]
    apres = contenu.split(FIN)[1]
    RECETTES_JS.write_text(avant + genere + apres, encoding="utf-8")

    print(f"OK — {len(blocs)} recettes écrites dans {RECETTES_JS.relative_to(RACINE)}.")


if __name__ == "__main__":
    main()
