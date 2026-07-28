# -*- coding: utf-8 -*-
"""
IMPORTER LES STATS DES PIOCHES  (Excel  ->  jeu/data/items.js)
=============================================================

Brioche règle les pioches de minage dans le classeur — la SOURCE :

    docs/BRUTAL-items-et-cartes.xlsx  ->  onglet « Pioches »

Puis on lance ce script pour régénérer la table lue par le jeu :

    python3 outils/importer_pioches.py

Le script réécrit UNIQUEMENT le bloc balisé de jeu/data/items.js :

    // <<PIOCHES-AUTO>>   ...généré...   // <<FIN-PIOCHES-AUTO>>

Il n'écrit que les trois chiffres d'ÉQUILIBRAGE — rareté, vitesse de minage,
efficacité. Le reste (taille dans le sac, icône, catégorie) reste écrit à la main
dans ITEMS : c'est de l'art et de la technique, pas de l'équilibrage.

Format de l'onglet : un en-tête, puis une ligne par pioche —
    A = id · B = Nom (info) · C = Rareté · D = Vitesse minage (%) · E = Efficacité (%)
    F = « mine jusqu'à » (info, déduit de la rareté) · G = note libre

RAPPEL DES RÈGLES (cf. jeu/systems/minage.js) :
  · efficacité 0 = exactement 1 minerai par coup ; 100 = un 2e garanti ;
    150 = un 2e garanti et 50 % d'en avoir un 3e ;
  · une pioche casse les minerais de SA rareté et d'UNE au-dessus.
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant :  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RACINE = Path(__file__).resolve().parent.parent
XLSX = RACINE / "docs" / "BRUTAL-items-et-cartes.xlsx"
ITEMS_JS = RACINE / "jeu" / "data" / "items.js"

DEBUT = "// <<PIOCHES-AUTO>>"
FIN = "// <<FIN-PIOCHES-AUTO>>"

RARETES_OK = {"commun", "uncommon", "rare", "epique", "legendaire"}


def lire_pioches(wb):
    if "Pioches" not in wb.sheetnames:
        raise SystemExit("L'onglet « Pioches » est introuvable dans le classeur.")
    ws = wb["Pioches"]
    pioches, vus = [], set()
    for i, lg in enumerate(ws.iter_rows(values_only=True), start=1):
        if not lg or not lg[0]:
            continue
        iid = str(lg[0]).strip()
        # On ne garde que les vraies lignes de données : un id en minuscules-tirets.
        # (« id » tout court est l'EN-TÊTE de colonne — il passe le motif, d'où le
        # test explicite : sans lui, l'importeur tentait de lire la ligne de titres.)
        if iid == "id" or not re.fullmatch(r"[a-z0-9-]+", iid):
            continue
        if iid in vus:
            raise SystemExit(f"« Pioches » ligne {i} : id en double « {iid} ».")
        vus.add(iid)

        rarete = str(lg[2]).strip() if len(lg) > 2 and lg[2] else ""
        if rarete not in RARETES_OK:
            raise SystemExit(f"« Pioches » ligne {i} : rareté « {rarete} » inconnue pour « {iid} » "
                             f"(attendues : {', '.join(sorted(RARETES_OK))}).")
        try:
            vitesse = int(round(float(lg[3])))
            efficacite = int(round(float(lg[4])))
        except (TypeError, ValueError, IndexError):
            raise SystemExit(f"« Pioches » ligne {i} : vitesse ou efficacité illisible pour « {iid} ».")
        if vitesse < 10:
            raise SystemExit(f"« Pioches » ligne {i} : vitesse {vitesse} trop basse pour « {iid} » "
                             "(minimum 10, sinon le minage devient interminable).")
        if efficacite < 0:
            raise SystemExit(f"« Pioches » ligne {i} : efficacité négative pour « {iid} ».")
        pioches.append((iid, rarete, vitesse, efficacite))
    if not pioches:
        raise SystemExit("Aucune pioche trouvée dans l'onglet « Pioches ».")
    return pioches


def bloc_js(pioches):
    lignes = ["const STATS_PIOCHES = {"]
    for iid, rarete, vitesse, efficacite in pioches:
        lignes.append(f'  "{iid}": {{ rarete: "{rarete}", vitesseMinage: {vitesse}, efficacite: {efficacite} }},')
    lignes.append("};")
    return "\n".join(lignes) + "\n"


def main():
    wb = load_workbook(XLSX, data_only=True)
    pioches = lire_pioches(wb)
    contenu = ITEMS_JS.read_text(encoding="utf-8")
    if DEBUT not in contenu or FIN not in contenu:
        raise SystemExit(f"Balises {DEBUT} / {FIN} introuvables dans {ITEMS_JS.name}.")
    avant = contenu.split(DEBUT)[0]
    apres = contenu.split(FIN)[1]
    ITEMS_JS.write_text(avant + DEBUT + "\n" + bloc_js(pioches) + FIN + apres, encoding="utf-8")
    print(f"OK — {len(pioches)} pioches écrites dans {ITEMS_JS.relative_to(RACINE)} : "
          + ", ".join(p[0] for p in pioches))


if __name__ == "__main__":
    main()
