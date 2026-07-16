#!/usr/bin/env python3
# Écrit des RÉSUMÉS D'ÉQUILIBRAGE dans le classeur, à partir du « Tableau de
# référence » de l'onglet « Lisez-moi » (édité par Brioche) :
#
#  • Onglet « Cartes » → colonne « Cible (barème) » : pour chaque carte, la valeur
#    à viser. Ex. une carte d'ATTAQUE Rare qui coûte 3 → 3×12 = 36 dégâts. Une carte
#    de DÉFENSE utilise « Armure / énergie ». ×1,5 si la carte ne vient QUE d'armes
#    à deux mains. Buffs et deck de base = « hors barème ».
#
#  • Onglet « Items » → colonne « Total visé (barème) » : pour chaque arme/armure,
#    le TOTAL que ses cartes devraient fournir (somme des cibles de chaque carte ×
#    quantité) : dégâts pour une arme, armure pour une pièce d'armure. Indique aussi
#    le nombre de cartes vs la cible de cartes de la rareté.
#
# Re-lançable à volonté (idempotent) : relance-le après avoir changé le tableau de
# référence pour tout rafraîchir.  →  python3 outils/resumer_cartes.py
#
# Dépendances : node (pour lire la liste des armes 2 mains) + openpyxl.

import json, os, re, subprocess
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "docs", "BRUTAL-items-et-cartes.xlsx")
ENTETE_CIBLE = "Cible (barème)"
ENTETE_TOTAL = "Total visé (barème)"


def armes_deux_mains():
    """Noms (anglais) des armes à DEUX MAINS, lus dans le code réel (items.js)."""
    code = (
        'import { ITEMS } from "./jeu/data/items.js";'
        'const n = Object.values(ITEMS).filter(i => i.mains === 2).map(i => i.nom);'
        'process.stdout.write(JSON.stringify(n));'
    )
    out = subprocess.check_output(["node", "--input-type=module", "-e", code], cwd=ROOT)
    return set(json.loads(out))


def lire_bareme(ws):
    """Lit le tableau de référence : rareté → nb de cartes, dégât/énergie, armure/énergie."""
    entete = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Rareté" and ws.cell(r, 3).value and "gât" in str(ws.cell(r, 3).value):
            entete = r
            break
    if entete is None:
        raise SystemExit("Tableau de référence introuvable dans « Lisez-moi ».")
    bareme = {}
    r = entete + 1
    while r <= ws.max_row and ws.cell(r, 1).value in ("Common", "Uncommon", "Rare", "Epic", "Legendary"):
        bareme[ws.cell(r, 1).value] = {
            "nb": ws.cell(r, 2).value,
            "degat": ws.cell(r, 3).value,
            "armure": ws.cell(r, 4).value,
        }
        r += 1
    return bareme


def entetes(ws, cle):
    """Trouve la ligne d'en-tête (colonne A == `cle`) et l'index des colonnes par nom."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == cle:
            noms = {ws.cell(r, c).value: c for c in range(1, ws.max_column + 2) if ws.cell(r, c).value}
            return r, noms
    raise SystemExit(f"En-tête « {cle} » introuvable.")


def est_deux_mains(donnee_par, armes2m):
    """Vrai si la carte n'est fournie QUE par des armes à deux mains."""
    if not donnee_par or donnee_par.strip().startswith("—"):
        return False
    sources = [s.strip() for s in donnee_par.split(",") if s.strip()]
    return bool(sources) and all(s in armes2m for s in sources)


def num(v):
    """Formatage compact : 36 et non 36.0 ; 1,5 avec virgule décimale (français)."""
    if v == int(v):
        return str(int(v))
    return ("%g" % v).replace(".", ",")


def cible_num(cout, type_, rarete, deux_mains, bareme):
    """(genre, valeur) d'une carte : genre ∈ {'degat','armure',None}, valeur numérique ou None.
    None = hors barème (buff, gratuite, deck de base)."""
    if rarete not in bareme or not isinstance(cout, (int, float)) or cout == 0:
        return (None, None)
    mult = 1.5 if deux_mains else 1
    if type_ == "Attaque":
        return ("degat", cout * bareme[rarete]["degat"] * mult)
    if type_ == "Defense":
        return ("armure", cout * bareme[rarete]["armure"] * mult)
    return (None, None)


def cible_texte(cout, type_, rarete, deux_mains, bareme):
    """Texte-résumé d'une carte pour la colonne « Cible (barème) »."""
    if rarete not in bareme:
        return "deck de base — hors barème"
    if not isinstance(cout, (int, float)):
        return ""
    if cout == 0:
        return "gratuite (0 énergie) — hors barème"
    genre, val = cible_num(cout, type_, rarete, deux_mains, bareme)
    if genre is None:
        return "hors barème (buff)"
    base = bareme[rarete]["degat" if genre == "degat" else "armure"]
    suffixe = "×1,5" if deux_mains else ""
    unite = "dégâts" if genre == "degat" else "armure"
    return f"≈ {num(val)} {unite}  ({cout}×{num(base)}{suffixe})"


def resumer_cartes(ws, bareme, armes2m):
    """Écrit la colonne « Cible (barème) » et renvoie un index nom_carte → (genre, valeur)."""
    ligne, cols = entetes(ws, "Famille")
    col = cols.get(ENTETE_CIBLE) or (max(cols.values()) + 1)
    c = ws.cell(ligne, col, ENTETE_CIBLE)
    c.font = Font(bold=True)
    c.comment = Comment(
        "Cible d'équilibrage calculée automatiquement (outils/resumer_cartes.py) :\n"
        "coût × (dégât ou armure par énergie selon la rareté, tableau « Lisez-moi »),\n"
        "×1,5 pour une carte fournie uniquement par des armes à deux mains.\n"
        "Ajuste l'effet de la carte pour t'en approcher (un peu au-dessus / en dessous, au choix).",
        "BRUTAL",
    )
    ws.column_dimensions[ws.cell(1, col).column_letter].width = 26

    index, n = {}, 0
    for r in range(ligne + 1, ws.max_row + 1):
        if not ws.cell(r, cols["ID"]).value:
            continue  # séparateur « ▸ … » ou ligne vide
        cout = ws.cell(r, cols["Coût"]).value
        type_ = ws.cell(r, cols["Type"]).value
        rarete = ws.cell(r, cols["Rareté"]).value
        d2m = est_deux_mains(ws.cell(r, cols["Donnée par"]).value, armes2m)
        cell = ws.cell(r, col, cible_texte(cout, type_, rarete, d2m, bareme))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        nom = ws.cell(r, cols["Nom"]).value
        if nom:
            index[re.sub(r"\s+", " ", str(nom).strip().lower())] = cible_num(cout, type_, rarete, d2m, bareme)
        n += 1
    return n, index


TOKEN = re.compile(r"^(.*?)\s*[×xX]\s*(\d+)\s*$")


def resumer_items(ws, bareme, index):
    """Écrit la colonne « Total visé (barème) » : total dégâts/armure des cartes d'un item."""
    ligne, cols = entetes(ws, "Catégorie")
    col = cols.get(ENTETE_TOTAL) or (max(cols.values()) + 1)
    c = ws.cell(ligne, col, ENTETE_TOTAL)
    c.font = Font(bold=True)
    c.comment = Comment(
        "Total que les cartes de l'item devraient fournir (outils/resumer_cartes.py) :\n"
        "somme des cibles de chaque carte × sa quantité — dégâts pour une arme, armure\n"
        "pour une armure. Entre parenthèses : nb de cartes de l'item vs cible de la rareté.",
        "BRUTAL",
    )
    ws.column_dimensions[ws.cell(1, col).column_letter].width = 30

    n = 0
    for r in range(ligne + 1, ws.max_row + 1):
        if not ws.cell(r, cols["ID"]).value:
            continue  # séparateur « ▸ … » ou ligne vide
        rarete = ws.cell(r, cols["Rareté"]).value
        cartes = ws.cell(r, cols["Cartes (×qté)"]).value
        deg = arm = 0.0
        nb_cartes = buffs = inconnues = 0
        for tok in str(cartes or "").split(","):
            tok = tok.strip()
            if not tok or tok == "—":
                continue
            m = TOKEN.match(tok)
            if not m:
                inconnues += 1
                continue
            nom = re.sub(r"\s+", " ", m.group(1).strip().lower())
            qte = int(m.group(2))
            nb_cartes += qte
            if nom not in index:
                inconnues += 1
                continue
            genre, val = index[nom]
            if genre == "degat":
                deg += val * qte
            elif genre == "armure":
                arm += val * qte
            else:
                buffs += qte

        # Texte : total dégâts et/ou armure, puis nb de cartes vs cible de rareté.
        morceaux = []
        if deg:
            morceaux.append(f"≈ {num(deg)} dégâts")
        if arm:
            morceaux.append(f"≈ {num(arm)} armure")
        if not morceaux:
            morceaux.append("aucun dégât/armure")
        extra = []
        cible_nb = bareme.get(rarete, {}).get("nb")
        extra.append(f"{nb_cartes} carte{'s' if nb_cartes > 1 else ''}"
                     + (f"/{cible_nb} visées" if cible_nb else ""))
        if buffs:
            extra.append(f"{buffs} buff")
        if inconnues:
            extra.append(f"{inconnues} non listée{'s' if inconnues > 1 else ''}")
        cell = ws.cell(r, col, "  ·  ".join(morceaux) + f"  ({' · '.join(extra)})")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n += 1
    return n


def main():
    armes2m = armes_deux_mains()
    wb = openpyxl.load_workbook(XLSX)
    bareme = lire_bareme(wb["Lisez-moi"])
    n_cartes, index = resumer_cartes(wb["Cartes"], bareme, armes2m)
    n_items = resumer_items(wb["Items"], bareme, index)
    wb.save(XLSX)
    print(f"{n_cartes} cartes + {n_items} items résumés · barème : {', '.join(bareme)}.")


if __name__ == "__main__":
    main()
