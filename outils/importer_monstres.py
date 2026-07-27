# -*- coding: utf-8 -*-
"""
IMPORTER LES STATS DES MONSTRES  (Excel  ->  jeu/data/ennemis.js)
=================================================================

Les CHIFFRES d'équilibrage des monstres (vie, dégâts, XP, vitesse, actions
spéciales) sont éditables dans le classeur — la SOURCE :

    docs/BRUTAL-items-et-cartes.xlsx  ->  onglet « Monstres »

Puis on régénère le bloc lu par le jeu :

    python3 outils/importer_monstres.py

Le script réécrit UNIQUEMENT le bloc balisé de jeu/data/ennemis.js :

    // <<MONSTRES-AUTO>>   ...généré...   // <<FIN-MONSTRES-AUTO>>

Ces stats REMPLACENT (source de vérité) les valeurs de secours écrites dans
ENNEMIS. L'ART et la technique (planche, sprite, portrait, butin…) restent, eux,
dans le code : l'Excel ne sert qu'à RÉGLER des monstres existants.

Onglet « Monstres » (ligne 1 = titres, ligne 2 = aide, données ensuite) :
    A=id · B=nom · C=niveau · D=famille · E=pv · F=attaque · G=xp · H=vitesse
    I=actions  (spéciales, format « type:valeur:poids » séparés par « | »)
Types d'action reconnus (doivent matcher le moteur, cf. systems/combat.js) :
    attaque · soigner · haste-allie          (vide = attaque simple)
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant :  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RACINE = Path(__file__).resolve().parent.parent
XLSX = RACINE / "docs" / "BRUTAL-items-et-cartes.xlsx"
ENN_JS = RACINE / "jeu" / "data" / "ennemis.js"

DEBUT = "// <<MONSTRES-AUTO>>"
FIN = "// <<FIN-MONSTRES-AUTO>>"

ACTIONS_OK = {"attaque", "soigner", "haste-allie", "buff-allie", "bouclier-allie", "renforts"}

# RECONNAISSANCE DES ACTIONS. Brioche écrit en français courant, pas en code : on
# cherche donc des MOTS-CLÉS dans la ligne plutôt qu'un libellé exact. L'ordre compte
# (« bouclier alliés » avant « alliés »). Comparaison sans accents ni casse, fautes de
# frappe courantes incluses (« fore » pour « force », « hate » pour « hâte »).
MOTS_ACTION = [
    (("bouclier", "shield"), "bouclier-allie"),
    (("soin", "soigne", "heal"), "soigner"),
    (("renfort", "reinforce"), "renforts"),
    # Un buff qui donne À LA FOIS de la force et de la hâte = le porte-étendard
    # (buff-allie fait déjà les deux : +hâte et +dégâts, cf. systems/combat.js).
    (("force", "fore", "degat", "etendard", "banniere", "buff"), "buff-allie"),
    (("hate", "haste", "celerite"), "haste-allie"),
    (("attaque", "attack", "frappe"), "attaque"),
]


def sans_accents(s):
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def echapper(s):
    return str(s).replace("\\", "\\\\").replace('"', '\\"')


def lire_attaque(cellule, iid, ligne):
    """Colonne « attaque ». TOLÈRE « NxM » (N dégâts × M coups, ex. « 2x2 ») → (N, M),
    et « N » simple → (N, 1). Le nombre de coups (M) alimente `attaqueHits`."""
    if isinstance(cellule, (int, float)):
        return int(round(cellule)), 1
    s = str(cellule) if cellule is not None else ""
    m = re.search(r"(-?\d+)\s*[x×X]\s*(\d+)", s)
    if m:
        return int(m.group(1)), max(1, int(m.group(2)))
    m2 = re.search(r"-?\d+", s)
    if not m2:
        raise SystemExit(f"« Monstres » ligne {ligne} ({iid}) : attaque illisible « {cellule} ».")
    return int(m2.group()), 1


def lire_int(cellule, iid, ligne, quoi):
    """Lit un entier, en TOLÉRANT une note humaine autour (ex. « 2x2 » → 2, « 5 (acide) »
    → 5). On prend le PREMIER entier trouvé. Sert à pv/attaque/xp/vitesse/niveau."""
    if isinstance(cellule, (int, float)):
        return int(round(cellule))
    m = re.search(r"-?\d+", str(cellule) if cellule is not None else "")
    if not m:
        raise SystemExit(f"« Monstres » ligne {ligne} ({iid}) : {quoi} illisible « {cellule} ».")
    return int(m.group())


def lire_actions(cellule, iid, ligne, attaqueBase=0):
    """Parse les ACTIONS SPÉCIALES (une par ligne).

    Une ligne est une ACTION si — et seulement si — elle porte une chance entre
    parenthèses : « … (60%) ». Tout le reste est une note libre, ignorée.
    Le TYPE vient d'un mot-clé (cf. MOTS_ACTION), la VALEUR est le premier nombre
    de la ligne (hors pourcentage) ; sans nombre, on prend l'attaque de base du
    monstre (« Attaque (30%) »).

    ⚠ Une ligne qui porte un « (n%) » mais dont on ne reconnaît pas le type est une
    ERREUR, pas une note. C'est ce silence-là qui avait fait disparaître le bouclier
    du Gobelin défenseur et le buff du Porte-étendard sans que personne le voie.
    """
    txt = str(cellule).strip() if cellule else ""
    if not txt:
        return []
    out = []
    for ln in re.split(r"[\n|]+", txt):
        ln = ln.strip()
        if not ln:
            continue
        # Ancien format « type:valeur:poids » (toujours accepté).
        if ":" in ln and ln.split(":")[0].strip() in ACTIONS_OK:
            bouts = [b.strip() for b in ln.split(":")]
            if len(bouts) != 3:
                raise SystemExit(f"« Monstres » ligne {ligne} ({iid}) : action « {ln} » mal formée.")
            typ, val, poids = bouts
            out.append((typ, int(round(float(val))), int(round(float(poids)))))
            continue
        # Format lisible : la CHANCE en fin de ligne, « (60%) ».
        mp = re.search(r"\(\s*(\d+)\s*%\s*\)\s*$", ln)
        if not mp:
            continue  # pas de pourcentage → note libre
        poids = int(mp.group(1))
        avant = ln[:mp.start()]
        plat = sans_accents(avant).lower()
        typ = next((t for mots, t in MOTS_ACTION if any(m in plat for m in mots)), None)
        if not typ:
            raise SystemExit(
                f"« Monstres » ligne {ligne} ({iid}) : action « {ln.strip()} » non comprise.\n"
                f"  Elle porte une chance ({poids}%) mais aucun mot-clé connu. Attendus : "
                f"bouclier, soin, renfort, force/étendard, hâte, attaque.")
        mv = re.search(r"-?\d+", avant)
        # Sans nombre : « Attaque (30%) » = l'attaque de base du monstre ; pour tout
        # le reste (ex. « Renforts (10%) ») la valeur n'a pas de sens → 0.
        val = int(mv.group()) if mv else (int(attaqueBase) if typ == "attaque" else 0)
        out.append((typ, val, poids))
    return out


def lire_monstres(wb):
    if "Monstres" not in wb.sheetnames:
        raise SystemExit("L'onglet « Monstres » est introuvable dans le classeur.")
    ws = wb["Monstres"]
    monstres, vus = [], set()
    for i, lg in enumerate(ws.iter_rows(values_only=True), start=1):
        if not lg:
            continue
        iid = str(lg[0]).strip() if lg[0] else ""
        if iid == "id" or not re.fullmatch(r"[a-z0-9-]+", iid):
            continue  # titre / aide / ligne vide
        if iid in vus:
            raise SystemExit(f"« Monstres » : id en double « {iid} ».")
        vus.add(iid)

        def champ(k):
            return lg[k] if len(lg) > k else None

        nom = str(champ(1)).strip() if champ(1) else iid
        niveau = lire_int(champ(2), iid, i, "niveau") if champ(2) is not None else 1
        famille = str(champ(3)).strip() if champ(3) else ""
        pv = lire_int(champ(4), iid, i, "pv")
        attaque, hits = lire_attaque(champ(5), iid, i)   # « 2x2 » → attaque 2, hits 2
        xp = lire_int(champ(6), iid, i, "xp")
        vitesse = lire_int(champ(7), iid, i, "vitesse")
        actions = lire_actions(champ(8), iid, i, attaque)
        # Colonne J = « grand » : 1/oui/true → monstre GRAND (occupe 2 places, rare).
        gv = str(champ(9)).strip().lower() if champ(9) is not None else ""
        grand = gv in ("1", "oui", "true", "vrai", "x")
        monstres.append(dict(id=iid, nom=nom, niveau=niveau, famille=famille, pv=pv,
                             attaque=attaque, hits=hits, xp=xp, vitesse=vitesse,
                             actions=actions, grand=grand))
    if not monstres:
        raise SystemExit("Aucun monstre trouvé dans l'onglet « Monstres ».")
    return monstres


def bloc_js(monstres):
    lignes = ["const STATS_MONSTRES = {"]
    for m in monstres:
        champs = [
            f'nom: "{echapper(m["nom"])}"',
            f'niveau: {m["niveau"]}',
        ]
        if m["famille"]:
            champs.append(f'famille: "{echapper(m["famille"])}"')
        champs += [
            f'pv: {m["pv"]}', f'attaque: {m["attaque"]}',
            f'xp: {m["xp"]}', f'vitesse: {m["vitesse"]}',
        ]
        if m.get("hits", 1) > 1:
            champs.append(f'attaqueHits: {m["hits"]}')
        if m["actions"]:
            acts = ", ".join(
                f'{{ type: "{t}", valeur: {v}, poids: {p} }}' for t, v, p in m["actions"])
            champs.append(f'actions: [{acts}]')
        else:
            champs.append("actions: []")
        if m["grand"]:
            champs.append("grand: true")
        lignes.append(f'  "{m["id"]}": {{ ' + ", ".join(champs) + " },")
    lignes.append("};")
    return "\n".join(lignes) + "\n"


def main():
    wb = load_workbook(XLSX, data_only=True)
    monstres = lire_monstres(wb)
    corps = bloc_js(monstres)

    contenu = ENN_JS.read_text(encoding="utf-8")
    if DEBUT not in contenu or FIN not in contenu:
        raise SystemExit(f"Balises {DEBUT} / {FIN} absentes de {ENN_JS.name}.")
    avant = contenu.split(DEBUT)[0]
    apres = contenu.split(FIN)[1]
    ENN_JS.write_text(f"{avant}{DEBUT}\n{corps}{FIN}{apres}", encoding="utf-8")
    print(f"OK — {len(monstres)} monstres écrits dans {ENN_JS.relative_to(RACINE)} : "
          + ", ".join(m["id"] for m in monstres))


if __name__ == "__main__":
    main()
