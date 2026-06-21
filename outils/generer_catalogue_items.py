#!/usr/bin/env python3
# Génère un classeur Excel listant tous les items du jeu (rangés par type, avec
# les cartes qu'ils ajoutent et leur quantité) + une feuille de référence des
# cartes existantes. Sert de support pour concevoir de nouvelles armes.
#
# Les données sont LUES DEPUIS LE CODE RÉEL (jeu/data/items.js + cartes.js) via
# Node, donc le fichier reste fidèle au jeu. Régénérer : python3 outils/generer_catalogue_items.py
#
# Dépendances : node (déjà dans le projet) + openpyxl (pip install openpyxl).

import json, os, subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SORTIE = os.path.join(ROOT, "docs", "BRUTAL-items-et-cartes.xlsx")

MAX_CARTES = 4   # nb de colonnes « Carte / Qté » (marge au-dessus du max actuel = 3)
LIGNES_VIDES = 6 # cases vides offertes sous chaque type pour ajouter des items


def extraire_donnees():
    """Importe ITEMS / RARETES / CARTES / SETS depuis les modules JS et renvoie un dict."""
    code = (
        'import { ITEMS, RARETES, SETS } from "./jeu/data/items.js";'
        'import { CARTES } from "./jeu/data/cartes.js";'
        'process.stdout.write(JSON.stringify({ITEMS, RARETES, CARTES, SETS}));'
    )
    out = subprocess.check_output(["node", "--input-type=module", "-e", code], cwd=ROOT)
    return json.loads(out)


# Libellé lisible d'un déclencheur de bonus de set.
DECLENCHEURS = {
    "frappeMelee": "Quand le héros est frappé en mêlée",
    "debutCombat": "En début de combat",
    "saignementCombo": "Quand un combo de saignement inflige des dégâts bonus",
}


def effets_hors_carte(item, set_par_item):
    """Texte décrivant ce qu'un item apporte EN DEHORS de ses cartes (armure de
    départ, passif propre, sac, deux mains, appartenance à un set). Vide si rien."""
    parts = []
    if item.get("mains") == 2:
        parts.append("Deux mains")
    if item.get("armureDepart"):
        parts.append(f"Ajoute {item['armureDepart']} d'armure en début de combat")
    if item.get("celeritePct") or item.get("vitesseBonus"):
        parts.append(f"+{item.get('celeritePct', 0)}% célérité, +{item.get('vitesseBonus', 0)} move speed")
    if item.get("passifPropre", {}).get("texte"):
        parts.append(item["passifPropre"]["texte"])
    if item.get("rangsBonus"):
        parts.append(f"+{item['rangsBonus']} rangées de sac")
    nom_set = set_par_item.get(item["id"])
    if nom_set:
        parts.append(f"Set : {nom_set}")
    return "  ·  ".join(parts)


# ---- Styles partagés --------------------------------------------------------
def fill(hexa):
    return PatternFill("solid", fgColor=hexa.lstrip("#").upper())

GRIS_BORD = Side(style="thin", color="C9C9C9")
BORD = Border(left=GRIS_BORD, right=GRIS_BORD, top=GRIS_BORD, bottom=GRIS_BORD)
F_TITRE   = Font(bold=True, size=15, color="FFFFFF")
F_NOTE    = Font(italic=True, size=10, color="555555")
F_ENTETE  = Font(bold=True, size=10, color="FFFFFF")
F_SECTION = Font(bold=True, size=12, color="F0E4B0")
F_GRAS    = Font(bold=True)
CENTRE = Alignment(horizontal="center", vertical="center")
GAUCHE = Alignment(horizontal="left", vertical="center", wrap_text=True)
FOND_ENTETE  = fill("23202E")
FOND_SECTION = fill("2E2640")
FOND_TITRE   = fill("7C2D12")
FOND_VIDE    = fill("FBF8F2")
FOND_NOUVEAU = fill("FFF3C4")  # surlignage des items/cartes marqués `nouveau` (revue facile)

# Couleur de fond d'une cartouche de type de carte (= couleurs du jeu).
TYPE_CARTE = {
    "attaque": ("Attaque", "7A3B32"),
    "defense": ("Défense", "2F5066"),
    "buff":    ("Buff",    "3F7A4D"),
}

GROUPES = [
    ("Armes — une main",      lambda it: it["categorie"] == "arme" and it.get("mains") == 1),
    ("Armes — deux mains",    lambda it: it["categorie"] == "arme" and it.get("mains") == 2),
    ("Main seconde / Off-hand (boucliers, grimoires, sceptres, torches)", lambda it: it["categorie"] == "bouclier"),
    ("Armures",               lambda it: it["categorie"] == "armure"),
    ("Colliers",              lambda it: it["categorie"] == "collier"),
    ("Bagues",                lambda it: it["categorie"] == "bague"),
    ("Gants",                 lambda it: it["categorie"] == "gant"),
    ("Bottes",                lambda it: it["categorie"] == "botte"),
    ("Sacs",                  lambda it: it["categorie"] == "sac"),
]


def cartes_comptees(item):
    """[(id_carte, quantité)] dans l'ordre d'apparition."""
    res = []
    for cid in (item.get("cartes") or []):
        for i, (c, n) in enumerate(res):
            if c == cid:
                res[i] = (c, n + 1)
                break
        else:
            res.append((cid, 1))
    return res


def construire():
    data = extraire_donnees()
    ITEMS, RARETES, CARTES = data["ITEMS"], data["RARETES"], data["CARTES"]
    SETS = data.get("SETS", {})
    nom_carte = {cid: c["nom"] for cid, c in CARTES.items()}
    # Item -> nom du set auquel il appartient (pour la colonne « Effet hors-carte »).
    set_par_item = {}
    for s in SETS.values():
        for pid in s.get("pieces", []):
            set_par_item[pid] = s["nom"]

    wb = Workbook()

    # =================== Feuille 1 : Lisez-moi ===================
    ws = wb.active
    ws.title = "Lisez-moi"
    ws.sheet_properties.tabColor = "7C2D12"
    ws.column_dimensions["A"].width = 100
    lignes = [
        ("BRUTAL — Catalogue des items & cartes", F_TITRE, FOND_TITRE),
        ("", None, None),
        ("À quoi sert ce fichier", F_GRAS, None),
        ("Lister proprement tout l'équipement actuel (onglet « Items »), rangé par type, avec "
         "les cartes que chaque pièce ajoute au deck et leur quantité. Sers-toi des lignes vides "
         "sous chaque type pour me proposer de nouvelles armes / armures.", None, None),
        ("", None, None),
        ("Comment ajouter une arme", F_GRAS, None),
        ("1) Va dans l'onglet « Items », sous le bon type.  2) Remplis une ligne vide : Nom, Rareté, "
         "puis pour chaque carte son Nom et sa Quantité (colonnes Carte 1/Qté, Carte 2/Qté…).", None, None),
        ("3) Si une carte EXISTE déjà, prends son nom dans l'onglet « Cartes » (référence).  "
         "Si tu inventes une NOUVELLE carte, écris son nom dans une colonne Carte et décris son "
         "effet (dégâts, Pierre, Poison, coût en Chaleur…) dans la colonne « Notes ».", None, None),
        ("", None, None),
        ("Effets HORS carte (armure de départ, sac, deux mains, set d'armure)", F_GRAS, None),
        ("La colonne « Effet hors-carte » (onglet Items) décrit ce qu'un objet apporte EN DEHORS "
         "de ses cartes : armure (Pierre) donnée en début de combat, rangées de sac, arme à deux "
         "mains, passif propre, appartenance à un set. Pour un nouvel objet, écris ici son effet "
         "hors-carte (ex. « Ajoute 12 d'armure en début de combat », « +4 rangées de sac »).", None, None),
        ("Les SETS d'armure ont leur propre onglet « Sets » : un bonus s'active quand on porte TOUTES "
         "les pièces d'armure du set (l'arme ne compte pas). Propose un set sur une ligne vide.", None, None),
        ("", None, None),
        ("Important : une arme n'a PAS de « dégâts » propres", F_GRAS, None),
        ("Dans BRUTAL, tout le combat passe par les CARTES du deck — une arme/armure ne fait que "
         "fournir des cartes. C'est pourquoi ce tableau se concentre sur les cartes et leur quantité, "
         "pas sur un chiffre de dégâts.", None, None),
        ("", None, None),
        ("Rareté des cartes (onglet « Cartes »)", F_GRAS, None),
        ("La couleur de rareté d'une carte = celle de l'objet LE MOINS rare qui la débloque "
         "(ex. « Strike » est Common car des armes communes la donnent). Calculée automatiquement "
         "à la génération — c'est PUREMENT indicatif pour s'y retrouver, cette rareté n'existe pas "
         "dans le jeu. « — » = carte du deck de base (donnée par aucun objet).", None, None),
        ("", None, None),
        ("Lignes surlignées en JAUNE", F_GRAS, None),
        ("Les items et cartes récemment ajoutés (à relire/valider) sont surlignés en jaune "
         "dans les onglets « Items » et « Cartes ». Une fois que tu les as validés, dis-le moi "
         "et je retire le surlignage.", None, None),
        ("", None, None),
        ("Légende des raretés", F_GRAS, None),
    ]
    r = 1
    for txt, font, bg in lignes:
        c = ws.cell(r, 1, txt)
        if font: c.font = font
        if bg: c.fill = bg
        c.alignment = GAUCHE
        if bg == FOND_TITRE:
            ws.row_dimensions[r].height = 24
        r += 1
    # pastilles de rareté
    for cle in ["commun", "uncommon", "rare", "epique", "legendaire"]:
        rr = RARETES[cle]
        c = ws.cell(r, 1, f"  {rr['nom']}  ({cle})")
        c.fill = fill(rr["couleur"])
        c.font = Font(bold=True, color=("000000" if cle == "commun" else "FFFFFF"))
        c.alignment = GAUCHE
        r += 1

    # Tableau de référence d'équilibrage (repère de design, fourni par Brioche).
    r += 1
    ws.cell(r, 1, "Tableau de référence pour garder une certaine balance").font = F_GRAS
    r += 1
    bal_entetes = ["Rareté", "Nb de cartes", "Dégât / énergie", "Armure / énergie"]
    for j, txt in enumerate(bal_entetes):
        c = ws.cell(r, 1 + j, txt); c.font = F_GRAS; c.alignment = GAUCHE
    r += 1
    bal_lignes = [
        ("Common", 5, 6, 5),
        ("Uncommon", 6, 9, 8),
        ("Rare", 6, 12, 10),
        ("Epic", 7, 16, 14),
        ("Legendary", 8, 20, 18),
    ]
    for nom_r, nb, deg, arm in bal_lignes:
        for j, val in enumerate([nom_r, nb, deg, arm]):
            ws.cell(r, 1 + j, val).alignment = GAUCHE
        r += 1
    ws.cell(r, 1, "Armes à deux mains : ×1,5 sur dégât et armure/énergie (moins de cartes), "
                  "souvent via des sorts chers en énergie.").alignment = GAUCHE
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 18

    # =================== Feuille 2 : Items ===================
    ws = wb.create_sheet("Items")
    ws.sheet_properties.tabColor = "2E2640"
    entetes = (["Nom", "ID", "Rareté", "Nb cartes"]
               + sum([[f"Carte {i}", "Qté"] for i in range(1, MAX_CARTES + 1)], [])
               + ["Effet hors-carte", "Notes / nouvelles cartes"])
    ncol = len(entetes)
    last = chr(ord("A") + ncol - 1)
    col_hors_carte = ncol - 1  # avant-dernière colonne (« Effet hors-carte »)

    # Titre + note
    ws.merge_cells(f"A1:{last}1")
    t = ws.cell(1, 1, "BRUTAL — Items par type (cartes ajoutées au deck)")
    t.font = F_TITRE; t.fill = FOND_TITRE; t.alignment = CENTRE
    ws.row_dimensions[1].height = 24
    ws.merge_cells(f"A2:{last}2")
    ws.cell(2, 1, "Une ligne = une pièce d'équipement. Remplis les lignes vides pour proposer de nouveaux items.").font = F_NOTE

    # En-têtes (ligne 3, figée)
    for j, h in enumerate(entetes, 1):
        c = ws.cell(3, j, h)
        c.font = F_ENTETE; c.fill = FOND_ENTETE; c.alignment = CENTRE; c.border = BORD
    ws.freeze_panes = "A4"

    # Largeurs
    larg = {"A": 24, "B": 21, "C": 12, "D": 9}
    for col, w in larg.items():
        ws.column_dimensions[col].width = w
    for i in range(MAX_CARTES):
        ws.column_dimensions[chr(ord("E") + i * 2)].width = 19   # Carte i
        ws.column_dimensions[chr(ord("E") + i * 2 + 1)].width = 6  # Qté i
    ws.column_dimensions[chr(ord("A") + col_hors_carte - 1)].width = 26  # Effet hors-carte
    ws.column_dimensions[last].width = 40                                # Notes

    row = 4
    for titre, test in GROUPES:
        membres = sorted(
            [it for it in ITEMS.values() if test(it)],
            key=lambda it: RARETES.get(it.get("rarete", "commun"), {}).get("rang", 0)
        )
        # ligne de section
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        c = ws.cell(row, 1, f"{titre}   ({len(membres)})")
        c.font = F_SECTION; c.fill = FOND_SECTION; c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1
        # items
        for it in membres:
            rr = RARETES.get(it["rarete"], {"nom": it["rarete"], "couleur": "9AA0A6"})
            cc = cartes_comptees(it)
            ws.cell(row, 1, it["nom"]).font = F_GRAS
            ws.cell(row, 2, it["id"]).font = Font(size=9, color="888888")
            cr = ws.cell(row, 3, rr["nom"])
            cr.fill = fill(rr["couleur"]); cr.alignment = CENTRE
            cr.font = Font(bold=True, color=("000000" if it["rarete"] == "commun" else "FFFFFF"))
            ws.cell(row, 4, len(it.get("cartes") or [])).alignment = CENTRE
            for i in range(MAX_CARTES):
                col = 5 + i * 2
                if i < len(cc):
                    cid, qte = cc[i]
                    ws.cell(row, col, nom_carte.get(cid, cid))
                    ws.cell(row, col + 1, qte).alignment = CENTRE
            ws.cell(row, col_hors_carte, effets_hors_carte(it, set_par_item)).font = F_NOTE
            for j in range(1, ncol + 1):
                ws.cell(row, j).border = BORD
                if ws.cell(row, j).alignment.horizontal is None:
                    ws.cell(row, j).alignment = GAUCHE
                # Surlignage des NOUVEAUX items (sauf la pastille de rareté, col 3).
                if it.get("nouveau") and j != 3:
                    ws.cell(row, j).fill = FOND_NOUVEAU
            row += 1
        # lignes vides à remplir
        for _ in range(LIGNES_VIDES):
            for j in range(1, ncol + 1):
                cell = ws.cell(row, j, None)
                cell.border = BORD; cell.fill = FOND_VIDE
            row += 1
        row += 1  # respiration entre sections

    # =================== Feuille 3 : Cartes (référence) ===================
    ws = wb.create_sheet("Cartes")
    ws.sheet_properties.tabColor = "3F7A4D"
    # qui donne quelle carte
    donne_par = {}
    for it in ITEMS.values():
        for cid in set(it.get("cartes") or []):
            donne_par.setdefault(cid, []).append(it["nom"])

    # Rareté d'une carte = celle de l'objet LE MOINS rare qui la débloque (rang
    # minimal). Ex. « Strike », donnée par des armes communes → Common. None si
    # aucun objet ne la donne (cartes du deck de base). PUREMENT pour le catalogue :
    # cette rareté n'existe pas dans le jeu, elle aide juste à s'y retrouver ici.
    rang_de = {cle: RARETES[cle]["rang"] for cle in RARETES}
    rarete_par_carte = {}
    for it in ITEMS.values():
        for cid in set(it.get("cartes") or []):
            actuelle = rarete_par_carte.get(cid)
            if actuelle is None or rang_de[it["rarete"]] < rang_de[actuelle]:
                rarete_par_carte[cid] = it["rarete"]

    entetes2 = ["Nom", "ID", "Coût (Chaleur)", "Type", "Rareté", "AOE", "Donnée par", "Effet"]
    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, "BRUTAL — Cartes existantes (référence pour composer des items)")
    t.font = F_TITRE; t.fill = FOND_TITRE; t.alignment = CENTRE
    ws.row_dimensions[1].height = 24
    for j, h in enumerate(entetes2, 1):
        c = ws.cell(2, j, h); c.font = F_ENTETE; c.fill = FOND_ENTETE; c.alignment = CENTRE; c.border = BORD
    ws.freeze_panes = "A3"
    for col, w in {"A": 18, "B": 21, "C": 13, "D": 11, "E": 12, "F": 6, "G": 28, "H": 62}.items():
        ws.column_dimensions[col].width = w

    ordre = {"attaque": 0, "defense": 1, "buff": 2}
    cartes_triees = sorted(CARTES.values(), key=lambda c: (ordre.get(c["type"], 9), c["nom"]))
    row = 3
    for cd in cartes_triees:
        ws.cell(row, 1, cd["nom"]).font = F_GRAS
        ws.cell(row, 2, cd["id"]).font = Font(size=9, color="888888")
        ws.cell(row, 3, cd["cout"]).alignment = CENTRE
        tnom, tcol = TYPE_CARTE.get(cd["type"], (cd["type"], "555555"))
        ct = ws.cell(row, 4, tnom); ct.fill = fill(tcol); ct.font = Font(bold=True, color="FFFFFF"); ct.alignment = CENTRE
        # Rareté (pastille colorée) — la moins rare des sources, ou « — » si deck de base.
        rcle = rarete_par_carte.get(cd["id"])
        if rcle:
            rr = RARETES[rcle]
            cr = ws.cell(row, 5, rr["nom"]); cr.fill = fill(rr["couleur"]); cr.alignment = CENTRE
            cr.font = Font(bold=True, color=("000000" if rcle == "commun" else "FFFFFF"))
        else:
            cr = ws.cell(row, 5, "—"); cr.alignment = CENTRE; cr.font = Font(color="888888")
        ws.cell(row, 6, "Oui" if cd.get("aoe") else "").alignment = CENTRE
        src = donne_par.get(cd["id"], [])
        if src:
            donne = ", ".join(src)
        elif cd["id"] in ("coup-faible", "garde-faible"):
            donne = "— (deck de base)"
        else:
            donne = "— (non équipée)"  # carte existante mais sur aucun objet actuellement
        ws.cell(row, 7, donne).alignment = GAUCHE
        ws.cell(row, 8, cd.get("texte", "")).alignment = GAUCHE
        for j in range(1, 9):
            ws.cell(row, j).border = BORD
            # Surlignage des NOUVELLES cartes (sauf type col 4 et rareté col 5, déjà colorées).
            if cd.get("nouveau") and j not in (4, 5):
                ws.cell(row, j).fill = FOND_NOUVEAU
        row += 1

    # =================== Feuille 4 : Sets d'armure ===================
    # Décrit les bonus de PANOPLIE (effets hors-carte qui s'activent quand toutes
    # les pièces d'armure d'un set sont équipées — l'arme ne compte pas).
    ws = wb.create_sheet("Sets")
    ws.sheet_properties.tabColor = "7C2D12"
    nom_item = {iid: it["nom"] for iid, it in ITEMS.items()}
    entetes3 = ["Set", "Pièces requises (armure seulement)", "Déclencheur", "Bonus"]
    ws.merge_cells("A1:D1")
    t = ws.cell(1, 1, "BRUTAL — Sets d'armure (bonus de panoplie, hors cartes)")
    t.font = F_TITRE; t.fill = FOND_TITRE; t.alignment = CENTRE
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:D2")
    ws.cell(2, 1, "Le bonus s'active quand TOUTES les pièces sont équipées en même temps "
                  "(l'arme n'est pas requise). Écris un nouveau set sur une ligne vide.").font = F_NOTE
    for j, h in enumerate(entetes3, 1):
        c = ws.cell(3, j, h); c.font = F_ENTETE; c.fill = FOND_ENTETE; c.alignment = CENTRE; c.border = BORD
    ws.freeze_panes = "A4"
    for col, w in {"A": 18, "B": 40, "C": 34, "D": 60}.items():
        ws.column_dimensions[col].width = w
    row = 4
    for s in SETS.values():
        pieces = "  +  ".join(nom_item.get(p, p) for p in s.get("pieces", []))
        bonus = s.get("bonus", {})
        decl = DECLENCHEURS.get(bonus.get("declencheur"), bonus.get("declencheur", ""))
        ws.cell(row, 1, s["nom"]).font = F_GRAS
        ws.cell(row, 2, pieces).alignment = GAUCHE
        ws.cell(row, 3, decl).alignment = GAUCHE
        ws.cell(row, 4, bonus.get("texte", "")).alignment = GAUCHE
        for j in range(1, 5):
            ws.cell(row, j).border = BORD
        row += 1
    # quelques lignes vides pour proposer de nouveaux sets
    for _ in range(LIGNES_VIDES):
        for j in range(1, 5):
            cell = ws.cell(row, j, None); cell.border = BORD; cell.fill = FOND_VIDE
        row += 1

    os.makedirs(os.path.dirname(SORTIE), exist_ok=True)
    wb.save(SORTIE)
    print(f"OK -> {SORTIE}  ({len(ITEMS)} items, {len(CARTES)} cartes, {len(SETS)} set(s))")


if __name__ == "__main__":
    construire()
