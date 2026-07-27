# -*- coding: utf-8 -*-
"""
IMPORTER LES STATS DE BASE DU HÉROS  (Excel  ->  jeu/data/heros_base.js)
=======================================================================

Les stats de base du héros (sans bonus ni talent) sont éditables dans le
classeur — la SOURCE :

    docs/BRUTAL-items-et-cartes.xlsx  ->  onglet « Héros »  (table clé / valeur)

Puis on régénère le bloc lu par le jeu :

    python3 outils/importer_heros.py

Le script réécrit UNIQUEMENT le bloc balisé de jeu/data/heros_base.js :

    // <<HEROS-AUTO>>   ...généré...   // <<FIN-HEROS-AUTO>>

Ces valeurs alimentent les constantes de base de systems/talents.js,
systems/combat.js et entities/heros.js.
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl manquant :  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RACINE = Path(__file__).resolve().parent.parent
XLSX = RACINE / "docs" / "BRUTAL-items-et-cartes.xlsx"
JS = RACINE / "jeu" / "data" / "heros_base.js"

DEBUT = "// <<HEROS-AUTO>>"
FIN = "// <<FIN-HEROS-AUTO>>"

# Clés attendues, DANS L'ORDRE de sortie. (Toutes obligatoires.)
CLES = ["pv", "vitesseMarche", "vitesseCombat", "chaleurDepart",
        "chaleurRecharge", "chaleurSeuil", "chaleurMax", "mainMax"]

# L'Excel parle la langue de Brioche, le code garde ses noms. « Agilité » y est le
# nom lisible de la vitesse d'initiative en combat. Comparaison sans accents ni casse.
SYNONYMES = {"agilite": "vitesseCombat", "agilité": "vitesseCombat"}


def normaliser(cle):
    import unicodedata
    sans = "".join(c for c in unicodedata.normalize("NFD", cle)
                   if unicodedata.category(c) != "Mn").lower()
    return SYNONYMES.get(sans, cle)


def lire_stats(wb):
    if "Héros" not in wb.sheetnames:
        raise SystemExit("L'onglet « Héros » est introuvable dans le classeur.")
    ws = wb["Héros"]
    vals = {}
    for lg in ws.iter_rows(values_only=True):
        if not lg:
            continue
        cle = normaliser(str(lg[0]).strip() if lg[0] else "")
        if cle not in CLES:
            continue  # titre / clé inconnue / ligne vide
        try:
            vals[cle] = int(round(float(lg[1])))
        except (TypeError, ValueError, IndexError):
            raise SystemExit(f"« Héros » : valeur illisible pour « {cle} ».")
    manquants = [c for c in CLES if c not in vals]
    if manquants:
        raise SystemExit(f"« Héros » : clé(s) manquante(s) : {', '.join(manquants)}.")
    return vals


def bloc_js(vals):
    lignes = ["export const STATS_HEROS_BASE = {"]
    for c in CLES:
        lignes.append(f"  {c}: {vals[c]},")
    lignes.append("};")
    return "\n".join(lignes) + "\n"


def main():
    wb = load_workbook(XLSX, data_only=True)
    vals = lire_stats(wb)
    corps = bloc_js(vals)

    contenu = JS.read_text(encoding="utf-8")
    if DEBUT not in contenu or FIN not in contenu:
        raise SystemExit(f"Balises {DEBUT} / {FIN} absentes de {JS.name}.")
    avant = contenu.split(DEBUT)[0]
    apres = contenu.split(FIN)[1]
    JS.write_text(f"{avant}{DEBUT}\n{corps}{FIN}{apres}", encoding="utf-8")
    print(f"OK — stats de base du héros écrites dans {JS.relative_to(RACINE)} : {vals}")


if __name__ == "__main__":
    main()
